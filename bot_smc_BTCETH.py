import asyncio
import json
import logging
import os
from dataclasses import dataclass, asdict # <-- AÑADIR asdict
from datetime import datetime
from typing import Dict, Optional, List
import time # Asegúrate de que 'time' esté importado

# Importamos ccxt (¡Asegúrate de que esté!)
import ccxt

# Importamos la nueva librería
from bitunix import BitunixClient

# Importar rastreador de wallet
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from wallet_tracker import WalletTracker

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scipy.signal import argrelextrema
from telegram import Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.error import TelegramError

# --- Configuración del Logging (sin cambios) ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('smc_trading_bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class Position:
    """Estructura para almacenar el estado de una posición."""
    symbol: str
    direction: str
    size: float
    entry_price: float
    entry_time: datetime
    entry_idx: int            # Aunque ya no se use mucho para gestión, se mantiene por ahora
    stop_loss: float          # Será 0.0 para copy trades
    original_stop_loss: float # Será 0.0 para copy trades
    take_profit: float        # Será 0.0 para copy trades
    liquidation_price: float
    margin_used: float
    is_copy: bool = False     # <-- Asegúrate que esta línea esté presente
    

class SmartMoneyLiveBot:
    """
    Bot de trading SMC multi-símbolo con persistencia de balance.
    """

    def __init__(
        self,
        api_key: str,
        api_secret: str,
        telegram_token: str,
        telegram_chat_id: str,
        symbols: List[str] = ['BTCUSDT'],
        initial_balance: float = 1000.0,
    ) -> None:
        self.symbols = symbols
        self.timeframe = '15m' # Asegúrate que sea 15m como en el backtester
        self.candle_limit = 300
        self.refresh_seconds = 15

        # --- PARÁMETROS AJUSTADOS (para coincidir con el backtester) ---
        self.structure_lookback = 20
        self.risk_reward_ratio = 2.0         # CAMBIADO de 2.5 a 2
        self.leverage = 20                   # (Esto se sobrescribe abajo)
        self.risk_per_trade_pct = 0.05       # CAMBIADO de 0.02 a 0.05 (5% de riesgo)
        self.max_candles_in_trade = 24

        # --- NUEVOS PARÁMETROS (copiados del backtester) ---
        self.pool_lookback_bars = 192        # 192 velas de 15m (48 horas)
        self.equal_tol = 0.0003              # 0.03%
        self.min_rr = 1.5                    # R:R mínimo para usar un pool
        
        self.enable_macd_filter = True
        self.macd_fast = 12
        self.macd_slow = 26
        self.macd_signal = 9
        # --- FIN NUEVOS PARÁMETROS ---

        # Telegram
        self.telegram_bot = Bot(token=telegram_token)
        self.telegram_chat_id = telegram_chat_id
        
        # --- LÓGICA DE PERSISTENCIA DE BALANCE ---
        self.initial_balance = initial_balance
        self.balance = initial_balance
        self.trades: list[Dict] = []
        
        self.excel_filenames: Dict[str, str] = {}
        self.dfs: Dict[str, Optional[pd.DataFrame]] = {s: None for s in self.symbols}
        self.positions: Dict[str, Optional[Position]] = {s: None for s in self.symbols}
        
        # Archivo para guardar posiciones abiertas
        self.positions_persistence_file = 'live_positions_state.json'
        # Cargar posiciones si el bot se reinició
        self._load_persistent_positions()
      
        self.last_signal_times: Dict[str, Optional[pd.Timestamp]] = {s: None for s in self.symbols}
        self.ccxt_symbols: Dict[str, str] = {s: s.replace('USDT', '/USDT') for s in self.symbols}
        
        # Rastreador de wallet
        self.wallet_tracker: Optional[WalletTracker] = None
        
        # Sistema de comandos de Telegram
        self.last_update_id = 0
        self.command_handlers = {
            '/wallet': self.handle_wallet_command,
        }
        
        # Concurrencia y apalancamiento por símbolo (del bot en vivo)
        self.max_concurrent_open = 4
        self.leverage_per_symbol: Dict[str, int] = {
            'BTCUSDT': 30,
            'ETHUSDT': 15,
            'HYPEUSDT': 15,
            'SOLUSDT': 15,
        }
        
        # Configuración por símbolo (se cargará desde JSON)
        self.symbol_configs: Dict[str, Dict] = {}
        
        last_known_balance = None
        last_trade_time = pd.Timestamp(0, tz='UTC') 

        for symbol in self.symbols:
            filename = f"live_report_SMC_{symbol.replace('/', '_')}.xlsx"
            self.excel_filenames[symbol] = filename
            
            balance_from_file, time_from_file = self._initialize_excel(filename) 
            
            if balance_from_file is not None and time_from_file > last_trade_time:
                last_known_balance = balance_from_file
                last_trade_time = time_from_file
        
        if last_known_balance is not None:
            self.balance = last_known_balance
            logger.info(f"✅ Balance restaurado desde el último trade en Excel: ${self.balance:,.2f}")
        else:
            logger.info(f"ℹ️ Iniciando con balance de configuración (no se encontraron trades): ${self.balance:,.2f}")
        # --- FIN PERSISTENCIA ---

        # Conexión con BitunixClient (PARA OPERAR)
        logger.info("Conectando a Bitunix para operaciones...")
        self.client = BitunixClient(api_key=api_key, api_secret=api_secret)

        # Conexión ccxt (SOLO PARA DATOS)
        logger.info("Configurando ccxt (Binance Futures) para obtener datos históricos...")
        self.data_exchange = ccxt.binance({
            'enableRateLimit': True,
            'options': {'defaultType': 'future'}
        })
        
        self.is_running: bool = False

    # --- Métodos de Reporte en Excel ---
    
    def _initialize_excel(self, filename: str) -> (Optional[float], pd.Timestamp):
        """
        Comprueba si el archivo Excel existe. Si no, lo crea con cabeceras.
        Si existe, lee la última fila para obtener el balance más reciente.
        Devuelve: (ultimo_balance, timestamp_ultimo_trade)
        """
        min_timestamp = pd.Timestamp(0, tz='UTC')
        
        if not os.path.exists(filename):
            wb = Workbook()
            ws_trades = wb.active
            ws_trades.title = "Trades"
            headers = [
                'Fecha Entrada', 'Hora Entrada', 'Fecha Salida', 'Hora Salida', 'Dirección', # 1-5
                'Precio Entrada', 'Precio Salida', 'Stop Loss', 'Take Profit', 'Liquidación', # 6-10
                'Tamaño Posición', 'Margen Usado', 'P/L USD', 'Razón Salida', 'Balance'      # 11-15
            ]
            for col, header in enumerate(headers, 1):
                cell = ws_trades.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filename)
            logger.info(f"📊 Nuevo archivo Excel creado: {filename}")
            return None, min_timestamp
        else:
            logger.info(f"📊 Usando archivo Excel existente: {filename}")
            try:
                wb = load_workbook(filename)
                ws = wb["Trades"]
                
                if ws.max_row <= 1:
                    logger.info("   ↳ Archivo existente pero sin trades. Usando balance inicial.")
                    return None, min_timestamp

                last_balance = ws.cell(row=ws.max_row, column=15).value
                last_exit_date_str = ws.cell(row=ws.max_row, column=3).value
                last_exit_time_str = ws.cell(row=ws.max_row, column=4).value
                
                if not last_exit_date_str or not last_exit_time_str:
                        logger.info("   ↳ El último trade en el archivo aún está abierto. Buscando el anterior...")
                        return None, min_timestamp

                last_timestamp = pd.Timestamp(f"{last_exit_date_str} {last_exit_time_str}", tz='UTC')

                if isinstance(last_balance, (int, float)):
                    logger.info(f"   ↳ Último balance leído de Excel: ${last_balance:,.2f} (del {last_timestamp})")
                    return float(last_balance), last_timestamp
                else:
                    logger.warning(f"   ↳ No se pudo leer el último balance de la fila {ws.max_row}. Valor: {last_balance}")
                    return None, min_timestamp

            except Exception as e:
                logger.error(f"Error al leer el archivo Excel existente {filename}: {e}", exc_info=True)
                return None, min_timestamp

    def _save_trade_to_excel(self, trade: Dict) -> None:
        try:
            symbol = trade['symbol']
            filename = self.excel_filenames[symbol]
            
            wb = load_workbook(filename)
            ws = wb["Trades"]
            
            entry_time = trade['entry_time']
            exit_time = trade['exit_time']
            
            row_data = [
                entry_time.strftime('%Y-%m-%d'), entry_time.strftime('%H:%M:%S'),
                exit_time.strftime('%Y-%m-%d'), exit_time.strftime('%H:%M:%S'),
                trade['direction'], trade['entry_price'], trade['exit_price'],
                trade['stop_loss'], trade['take_profit'], trade['liquidation_price'],
                trade['position_size'], trade['margin_used'], trade['pnl'],
                trade['exit_reason'], self.balance
            ]
            ws.append(row_data)
            
            pnl_cell = ws.cell(row=ws.max_row, column=13)
            if trade['pnl'] > 0:
                pnl_cell.font = Font(color="00B050", bold=True)
            elif trade['pnl'] < 0:
                pnl_cell.font = Font(color="FF0000", bold=True)

            wb.save(filename)
            logger.info(f"📊 Trade de {symbol} guardado en {filename}.")
        except Exception as e:
            logger.error(f"Error al guardar en Excel: {e}", exc_info=True)

    # --- Métodos de Notificación por Telegram (sin cambios) ---
    async def send_telegram_message(self, message: str) -> None:
        try:
            await self.telegram_bot.send_message(chat_id=self.telegram_chat_id, text=message, parse_mode='HTML')
        except TelegramError as e:
            logger.error(f"Error al enviar mensaje de Telegram: {e}")
    
    async def listen_telegram_commands(self):
        """Escucha comandos de Telegram y callbacks de botones en background."""
        logger.info("🎧 Iniciando escucha de comandos de Telegram...")
        
        while self.is_running:
            try:
                # Obtener actualizaciones de Telegram
                updates = await self.telegram_bot.get_updates(
                    offset=self.last_update_id + 1,
                    timeout=10
                )
                
                for update in updates:
                    self.last_update_id = update.update_id
                    
                    # Verificar si es un callback query (botón presionado)
                    if update.callback_query:
                        callback = update.callback_query
                        chat_id = str(callback.message.chat_id)
                        
                        # Verificar que el callback viene del chat autorizado
                        if chat_id != self.telegram_chat_id:
                            logger.warning(f"Callback recibido de chat no autorizado: {chat_id}")
                            continue
                        
                        callback_data = callback.data
                        message_id = callback.message.message_id
                        
                        logger.info(f"🔘 Botón presionado: {callback_data}")
                        
                        # Responder al callback para quitar el "loading"
                        try:
                            await self.telegram_bot.answer_callback_query(callback.id)
                        except:
                            pass
                        
                        # Manejar diferentes callbacks
                        if callback_data == "refresh_wallet":
                            await self.handle_wallet_command(message_id=message_id)
                    
                    # Verificar si es un mensaje de texto
                    elif update.message and update.message.text:
                        text = update.message.text.strip()
                        chat_id = str(update.message.chat_id)
                        
                        # Verificar que el mensaje viene del chat autorizado
                        if chat_id != self.telegram_chat_id:
                            logger.warning(f"Comando recibido de chat no autorizado: {chat_id}")
                            continue
                        
                        logger.info(f"📨 Comando recibido: {text}")
                        
                        # Buscar handler para el comando
                        for command, handler in self.command_handlers.items():
                            if text.startswith(command):
                                try:
                                    await handler()
                                except Exception as e:
                                    logger.error(f"Error ejecutando comando {command}: {e}")
                                    await self.send_telegram_message(
                                        f"🚨 Error ejecutando comando {command}: {str(e)}"
                                    )
                                break
                
                await asyncio.sleep(2)  # Esperar 2 segundos entre polls
                
            except Exception as e:
                logger.error(f"Error en listener de comandos de Telegram: {e}")
                await asyncio.sleep(5)

    async def notify_entry(self, pos: Position) -> None:
        pnl_target = (pos.take_profit - pos.entry_price) * pos.size if pos.direction == 'LONG' else (pos.entry_price - pos.take_profit) * pos.size
        msg = (
            f"✅ <b>POSICIÓN ABIERTA (SMC)</b>\n\n"
            f"📊 Par: {pos.symbol}\n"
            f"📈 Dirección: <b>{pos.direction}</b>\n"
            f"💰 Precio entrada: ${pos.entry_price:,.4f}\n"
            f"📏 Tamaño: {pos.size:.4f}\n"
            f"💵 Margen: ${pos.margin_used:,.2f}\n"
            f"🎯 Take Profit: ${pos.take_profit:,.4f} (+${pnl_target:,.2f})\n"
            f"🛑 Stop Loss: ${pos.stop_loss:,.4f}\n"
            f"⚠️ Liquidación: ${pos.liquidation_price:,.4f}"
        )
        await self.send_telegram_message(msg)

    async def notify_exit(self, trade: Dict) -> None:
        emoji = "🟢" if trade['pnl'] > 0 else "🔴"
        msg = (
            f"{emoji} <b>POSICIÓN CERRADA (SMC)</b>\n\n"
            f"📊 Par: {trade['symbol']}\n"
            f"📈 Dirección: {trade['direction']}\n"
            f"💰 Entrada: ${trade['entry_price']:,.4f} | Salida: ${trade['exit_price']:,.4f}\n"
            f"💵 P/L: <b>${trade['pnl']:,.2f}</b>\n"
            f"📝 Razón: {trade['exit_reason']}\n"
            f"💼 Balance: ${self.balance:,.2f}"
        )
        await self.send_telegram_message(msg)

    # --- Metodos de Persistencia ---

    # --- Métodos de Persistencia ---

    def _load_persistent_positions(self) -> None:
        """Carga las posiciones abiertas desde un archivo JSON al iniciar."""
        if not os.path.exists(self.positions_persistence_file):
            logger.info("No se encontró archivo de estado de posiciones (live_positions_state.json). Empezando de cero.")
            return

        try:
            with open(self.positions_persistence_file, 'r') as f:
                persistent_data = json.load(f)
            
            reloaded_count = 0
            # Usamos list(persistent_data.items()) para evitar error si el dict cambia durante iteración
            for symbol, pos_data in list(persistent_data.items()): 
                if symbol in self.symbols and pos_data is not None:
                    # Asegurar que todos los campos esperados existan, con default para 'is_copy'
                    pos_data.setdefault('is_copy', False) # Añade is_copy=False si no existe (compatibilidad)
                    
                    # Convertir el string ISO de vuelta a datetime
                    pos_data['entry_time'] = datetime.fromisoformat(pos_data['entry_time'])
                    
                    # Recrear el objeto Position
                    try:
                        # Asegurar que los campos numéricos sean float (JSON puede cargarlos como int a veces)
                        pos_data['size'] = float(pos_data['size'])
                        pos_data['entry_price'] = float(pos_data['entry_price'])
                        pos_data['stop_loss'] = float(pos_data['stop_loss'])
                        pos_data['original_stop_loss'] = float(pos_data['original_stop_loss'])
                        pos_data['take_profit'] = float(pos_data['take_profit'])
                        pos_data['liquidation_price'] = float(pos_data['liquidation_price'])
                        pos_data['margin_used'] = float(pos_data['margin_used'])
                        
                        self.positions[symbol] = Position(**pos_data)
                        reloaded_count += 1
                        copy_status = "(COPY)" if pos_data['is_copy'] else "(SMC)"
                        logger.info(f"🔄 Posición persistente para {symbol} {copy_status} ({pos_data['direction']}) recargada.")
                    except TypeError as te:
                         logger.error(f"Error al recrear Position para {symbol} desde JSON: {te}. Datos: {pos_data}", exc_info=True)
                         # Si falla la recreación, mejor no cargarla para evitar errores posteriores
                         if symbol in self.positions:
                             self.positions[symbol] = None 
                    except KeyError as ke:
                         logger.error(f"Falta el campo '{ke}' en los datos JSON para {symbol}. Datos: {pos_data}. No se puede recargar.")
                         if symbol in self.positions:
                             self.positions[symbol] = None
                elif symbol not in self.symbols and pos_data is not None:
                     logger.warning(f"Se encontró posición persistente para {symbol} pero no está en la lista actual de símbolos ({self.symbols}). Ignorando.")

            if reloaded_count > 0:
                 logger.info(f"✅ Se recargaron {reloaded_count} posiciones abiertas.")

        except json.JSONDecodeError:
            logger.error(f"Error al decodificar JSON desde {self.positions_persistence_file}. El archivo podría estar corrupto o vacío.")
            if os.path.exists(self.positions_persistence_file):
                try:
                    os.remove(self.positions_persistence_file)
                    logger.info(f"Archivo JSON corrupto {self.positions_persistence_file} eliminado.")
                except OSError as e:
                    logger.error(f"No se pudo eliminar el archivo JSON corrupto: {e}")
        except Exception as e:
            logger.error(f"Error inesperado al cargar estado de posiciones desde {self.positions_persistence_file}: {e}", exc_info=True)
            logger.warning("Iniciando con posiciones vacías debido a un error de carga.")
            if os.path.exists(self.positions_persistence_file):
                 try:
                    os.remove(self.positions_persistence_file)
                    logger.info(f"Archivo JSON {self.positions_persistence_file} eliminado debido a error de carga.")
                 except OSError as e:
                    logger.error(f"No se pudo eliminar el archivo JSON tras error de carga: {e}")


    async def _save_persistent_positions(self) -> None:
        """Guarda el estado actual de self.positions en un archivo JSON."""
        logger.debug("Guardando estado de posiciones persistentes...")
        temp_file = self.positions_persistence_file + ".tmp" # Definir archivo temporal
        try:
            # Crear un diccionario serializable
            serializable_positions = {}
            for symbol, pos in self.positions.items():
                if pos is not None:
                    # Convertir objeto Position a dict
                    pos_dict = asdict(pos)
                    # Convertir datetime a string ISO 8601
                    # Asegurarse que entry_time es datetime antes de llamar isoformat
                    if isinstance(pos_dict['entry_time'], datetime):
                        pos_dict['entry_time'] = pos_dict['entry_time'].isoformat()
                    else:
                        # Si por alguna razón no es datetime, loguear error y no guardar esa pos
                        logger.error(f"Tipo inesperado para entry_time en posición {symbol}: {type(pos_dict['entry_time'])}. No se guardará esta posición.")
                        continue # Saltar esta posición al guardar
                    serializable_positions[symbol] = pos_dict
                else:
                    serializable_positions[symbol] = None # Guardar None para posiciones cerradas
            
            # Escribir a disco de forma segura (atomic write)
            with open(temp_file, 'w') as f:
                json.dump(serializable_positions, f, indent=4)
            # Renombrar el archivo temporal al final (operación atómica en la mayoría de S.O.)
            os.replace(temp_file, self.positions_persistence_file) 
            logger.debug(f"✅ Estado de posiciones guardado en {self.positions_persistence_file}.")

        except Exception as e:
            logger.error(f"Error crítico al guardar estado de posiciones en {self.positions_persistence_file}: {e}", exc_info=True)
            # Intentar eliminar archivo temporal si existe y falló el guardado
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                    logger.info(f"Archivo temporal {temp_file} eliminado tras error de guardado.")
                except OSError as ose:
                    logger.error(f"No se pudo eliminar archivo temporal {temp_file} tras error: {ose}")
    

    # --- Métodos de Obtención de Datos ---
    
    # <<< FUNCIÓN MODIFICADA >>>
    async def update_market_data(self, symbol: str) -> bool:
        """
        Actualiza los datos de mercado usando ccxt.fetch_ohlcv en chunks
        y calcula TODOS los indicadores y patrones SMC.
        """
        ccxt_symbol = self.ccxt_symbols[symbol]
        logger.info(f"🔄 [{symbol}] Actualizando datos de mercado ({self.candle_limit} velas)...")
        
        try:
            limit_per_request = 1000 
            all_ohlcv = []
            # Asegurar msec_per_candle dinámico basado en timeframe
            if 'm' in self.timeframe:
                msec_per_candle = int(self.timeframe.replace('m', '')) * 60 * 1000
            elif 'h' in self.timeframe:
                 msec_per_candle = int(self.timeframe.replace('h', '')) * 60 * 60 * 1000
            else:
                 msec_per_candle = 900000 # Default 15m
            
            msec_needed = self.candle_limit * 1.5 * msec_per_candle
            since = self.data_exchange.milliseconds() - int(msec_needed)

            while True:
                logger.debug(f"   [{symbol}] Descargando chunk desde timestamp {since}...")
                ohlcv = await asyncio.to_thread(
                    self.data_exchange.fetch_ohlcv,
                    ccxt_symbol, self.timeframe, since=since, limit=limit_per_request
                )
                if not ohlcv:
                    logger.info(f"   [{symbol}] No más datos disponibles.")
                    break
                all_ohlcv.extend(ohlcv)
                logger.debug(f"   Obtenidas {len(ohlcv)} velas (total: {len(all_ohlcv)})")
                since = ohlcv[-1][0] + 1 
                if len(all_ohlcv) >= self.candle_limit or len(ohlcv) < limit_per_request:
                    logger.debug(f"   [{symbol}] Alcanzado límite de velas ({len(all_ohlcv)}) o fin de datos.")
                    break
                await asyncio.sleep(self.data_exchange.rateLimit / 1000)

            if not all_ohlcv:
                logger.error(f"[{symbol}] No se pudieron descargar datos con ccxt.")
                return False

            df = pd.DataFrame(all_ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
            df = df.drop_duplicates(subset=['timestamp'], keep='last') 
            df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms', utc=True) 
            df.set_index('timestamp', inplace=True)
            df = df.sort_index()

            if len(df) > self.candle_limit:
                 df = df.tail(self.candle_limit)
            elif len(df) < self.structure_lookback + 2: 
                 logger.warning(f"[{symbol}] No se recibieron suficientes datos ({len(df)} velas).")
                 return False
            
            df = df.astype(float) 

            # --- INICIO MODIFICACIÓN: CÁLCULO DE PATRONES EN VELAS CERRADAS ---
            logger.debug(f"[{symbol}] Separando {len(df)-1} velas cerradas y 1 en desarrollo.")
            df_closed = df.iloc[:-1].copy()
            df_current = df.iloc[[-1]].copy()

            # 1. Calcular Estructura (min/max) SÓLO en velas cerradas
            n = self.structure_lookback
            if len(df_closed) > n:
                min_indices = argrelextrema(df_closed.low.values, np.less_equal, order=n)[0]
                max_indices = argrelextrema(df_closed.high.values, np.greater_equal, order=n)[0]
                df_closed['min'] = np.nan
                df_closed['max'] = np.nan
                
                valid_min_indices_loc = df_closed.index[min_indices[min_indices < len(df_closed)]]
                valid_max_indices_loc = df_closed.index[max_indices[max_indices < len(df_closed)]]
                if not valid_min_indices_loc.empty:
                    df_closed.loc[valid_min_indices_loc, 'min'] = df_closed.loc[valid_min_indices_loc, 'low']
                if not valid_max_indices_loc.empty:
                    df_closed.loc[valid_max_indices_loc, 'max'] = df_closed.loc[valid_max_indices_loc, 'high']
            else:
                 df_closed['min'] = np.nan
                 df_closed['max'] = np.nan

            # 2. Calcular FVGs SÓLO en velas cerradas
            df_closed['fvg_bull_high'], df_closed['fvg_bull_low'] = np.nan, np.nan
            df_closed['fvg_bear_high'], df_closed['fvg_bear_low'] = np.nan, np.nan
            for i in range(2, len(df_closed)): 
                 if df_closed['low'].iloc[i] > df_closed['high'].iloc[i-2]:
                     df_closed.loc[df_closed.index[i-1], 'fvg_bull_low'] = df_closed['high'].iloc[i-2]
                     df_closed.loc[df_closed.index[i-1], 'fvg_bull_high'] = df_closed['low'].iloc[i]
                 if df_closed['high'].iloc[i] < df_closed['low'].iloc[i-2]:
                     df_closed.loc[df_closed.index[i-1], 'fvg_bear_low'] = df_closed['high'].iloc[i]
                     df_closed.loc[df_closed.index[i-1], 'fvg_bear_high'] = df_closed['low'].iloc[i-2]
            
            # 3. Añadir NaNs a la vela actual
            df_current['min'], df_current['max'] = np.nan, np.nan
            df_current['fvg_bull_high'], df_current['fvg_bull_low'] = np.nan, np.nan
            df_current['fvg_bear_high'], df_current['fvg_bear_low'] = np.nan, np.nan
            
            # 4. Unir los DataFrames
            df_final = pd.concat([df_closed, df_current])
            
            # --- MODIFICADO: Calcular ATR y MACD en el DF final ---
            df_final = self.compute_atr(df_final)
            df_final = self.compute_macd(df_final)
            # --------------------------------------------------
            
            self.dfs[symbol] = df_final
            logger.info(f"✅ [{symbol}] Datos ({len(df_final)} velas) y patrones SMC (en {len(df_closed)} cerradas) actualizados.")
            return True

        except ccxt.NetworkError as e:
            logger.error(f"[{symbol}] Error de red ccxt: {e}")
            return False
        except ccxt.ExchangeError as e:
            logger.error(f"[{symbol}] Error del exchange ccxt: {e}")
            if isinstance(e, ccxt.RateLimitExceeded):
                logger.warning(f"[{symbol}] Rate limit excedido, esperando más tiempo...")
                await asyncio.sleep(60) 
            return False
        except Exception as e:
            logger.error(f"[{symbol}] Error inesperado actualizando datos: {e}", exc_info=True)
            return False

    # --- NUEVAS FUNCIONES (Copiadas del Backtester) ---

    def compute_atr(self, df: pd.DataFrame, window: int = 14) -> pd.DataFrame:
        """Calcula ATR simple y lo añade al DataFrame."""
        if df.empty: return df
        df_copy = df.copy() # Evitar SettingWithCopyWarning
        high = df_copy['high']
        low = df_copy['low']
        close = df_copy['close']
        prev_close = close.shift(1)
        tr = pd.concat([
            (high - low),
            (high - prev_close).abs(),
            (low - prev_close).abs()
        ], axis=1).max(axis=1)
        df_copy['atr'] = tr.rolling(window=window, min_periods=window).mean()
        return df_copy

    def compute_macd(self, df: pd.DataFrame, fast: int = None, slow: int = None, signal: int = None):
        """Calcula MACD clásico y lo añade al DataFrame."""
        if df.empty: return df
        df_copy = df.copy()
        if fast is None: fast = self.macd_fast
        if slow is None: slow = self.macd_slow
        if signal is None: signal = self.macd_signal
        
        close = df_copy['close']
        ema_fast = close.ewm(span=fast, adjust=False).mean()
        ema_slow = close.ewm(span=slow, adjust=False).mean()
        macd_line = ema_fast - ema_slow
        macd_signal_line = macd_line.ewm(span=signal, adjust=False).mean()
        macd_hist = macd_line - macd_signal_line
        
        df_copy['macd'] = macd_line
        df_copy['macd_signal'] = macd_signal_line
        df_copy['macd_hist'] = macd_hist
        return df_copy

    def _binsize(self, ref_price: float, tol: float) -> float:
        return max(1e-8, ref_price * tol)

    def build_liquidity_pools(self, df: pd.DataFrame, i: int, lookback: int = None, tol: float = None):
        """Construye pools de liquidez (ver backtester para detalles)."""
        if lookback is None: lookback = self.pool_lookback_bars
        if tol is None: tol = self.equal_tol
        start = max(0, i - lookback)
        window = df.iloc[start:i] # Usa velas cerradas hasta 'i' (excluye i)
        if window.empty: return []

        mid_price = float(window['close'].iloc[-1])
        binsize = self._binsize(mid_price, tol)
        step = 5.0 if mid_price < 5000 else 10.0 # Ajusta 'step' si operas otros activos
        pools: Dict[float, float] = {}

        def add(price: float, score: float):
            if price is None or np.isnan(price): return
            bucket = round(price / binsize)
            level = bucket * binsize
            pools[level] = pools.get(level, 0.0) + score

        # Equal highs/lows
        highs = window['high'].values; lows = window['low'].values
        for arr, base in ((highs, 3.0), (lows, 3.0)):
            counts: Dict[int, int] = {}
            for p in arr:
                b = round(p / binsize)
                counts[b] = counts.get(b, 0) + 1
            for b, cnt in counts.items():
                if cnt >= 2: add(b * binsize, base * cnt)

        # Swings
        if 'max' in window.columns:
            for p in window['max'].dropna().values: add(float(p), 4.0)
        if 'min' in window.columns:
            for p in window['min'].dropna().values: add(float(p), 4.0)
        
        # FVG borders
        if 'fvg_bull_high' in window.columns:
            for p in window['fvg_bull_high'].dropna().values: add(float(p), 2.5)
        if 'fvg_bear_low' in window.columns:
            for p in window['fvg_bear_low'].dropna().values: add(float(p), 2.5)

        # Niveles redondos
        wmin = float(window['low'].min()); wmax = float(window['high'].max())
        if step > 0: # Evitar bucle infinito si step es 0
            lvl = (np.floor(wmin / step) * step)
            while lvl <= wmax:
                hits = ((np.abs(window['high'] - lvl) <= binsize) | (np.abs(window['low'] - lvl) <= binsize)).sum()
                if hits >= 1: add(lvl, 0.5 * hits)
                lvl += step

        levels = [{'price': float(k), 'score': float(v)} for k, v in pools.items()]
        levels.sort(key=lambda x: (-x['score'], x['price']))
        return levels

    def select_target_pool(self, df: pd.DataFrame, i: int, direction: str, entry_price: float, pools: List[Dict]):
        """Elige el pool objetivo (ver backtester para detalles)."""
        if not pools: return None
        
        # Usamos ATR de la vela cerrada anterior (i-1) para la decisión
        atr_val = df['atr'].iloc[i-1] if 'atr' in df.columns and i-1 >= 0 and not pd.isna(df['atr'].iloc[i-1]) else np.nan
        max_dist = 1.5 * atr_val if not np.isnan(atr_val) else None

        if direction == 'LONG':
            candidates = [p for p in pools if p['price'] > entry_price]
            candidates.sort(key=lambda p: (-p['score'], abs(p['price'] - entry_price)))
            if max_dist is not None:
                within = [p for p in candidates if (p['price'] - entry_price) <= max_dist]
                if within: return within[0]['price']
            return candidates[0]['price'] if candidates else None
        else: # SHORT
            candidates = [p for p in pools if p['price'] < entry_price]
            candidates.sort(key=lambda p: (-p['score'], abs(p['price'] - entry_price)))
            if max_dist is not None:
                within = [p for p in candidates if (entry_price - p['price']) <= max_dist]
                if within: return within[0]['price']
            return candidates[0]['price'] if candidates else None

    # --- FIN NUEVAS FUNCIONES ---

    # --- Lógica de Setups (sin cambios) ---
    # Analiza el 'idx' (que ahora será -1) contra los patrones calculados
    def _atr_series(self, df: pd.DataFrame, window: int = 14) -> pd.Series:
        prev_close = df['close'].shift(1)
        tr = pd.concat([
            (df['high'] - df['low']),
            (df['high'] - prev_close).abs(),
            (df['low'] - prev_close).abs()
        ], axis=1).max(axis=1)
        return tr.rolling(window=window, min_periods=window).mean()

    def _binsize(self, ref_price: float, tol: float) -> float:
        return max(1e-8, ref_price * tol)

    def build_liquidity_pools(self, df: pd.DataFrame, i: int, lookback: int = None, tol: float = None):
        if lookback is None:
            lookback = self.pool_lookback_bars
        if tol is None:
            tol = self.equal_tol
        start = max(0, i - lookback)
        window = df.iloc[start:i]
        if window.empty:
            return []
        mid_price = float(window['close'].iloc[-1])
        binsize = self._binsize(mid_price, tol)
        step = 5.0 if mid_price < 5000 else 10.0
        pools: Dict[float, float] = {}

        def add(price: float, score: float):
            if price is None or np.isnan(price):
                return
            bucket = round(price / binsize)
            level = bucket * binsize
            pools[level] = pools.get(level, 0.0) + score

        highs = window['high'].values
        lows = window['low'].values
        for arr, base in ((highs, 3.0), (lows, 3.0)):
            counts: Dict[int, int] = {}
            for p in arr:
                b = round(p / binsize)
                counts[b] = counts.get(b, 0) + 1
            for b, cnt in counts.items():
                if cnt >= 2:
                    add(b * binsize, base * cnt)

        if 'max' in window.columns:
            for p in window['max'].dropna().values:
                add(float(p), 4.0)
        if 'min' in window.columns:
            for p in window['min'].dropna().values:
                add(float(p), 4.0)

        # FVG borders (si existen columnas de FVG en live df_closed; aquí usamos "borders" derivados)
        if 'fvg_bull_high' in window.columns and 'fvg_bear_low' in window.columns:
            # borde bull: fvg_bull_high (nivel de entrada long en tu lógica)
            for p in window['fvg_bull_high'].dropna().values:
                add(float(p), 2.5)
            # borde bear: fvg_bear_low (nivel de entrada short)
            for p in window['fvg_bear_low'].dropna().values:
                add(float(p), 2.5)

        wmin = float(window['low'].min())
        wmax = float(window['high'].max())
        lvl = (np.floor(wmin / step) * step)
        while lvl <= wmax:
            hits = ((np.abs(window['high'] - lvl) <= binsize) | (np.abs(window['low'] - lvl) <= binsize)).sum()
            if hits >= 1:
                add(lvl, 0.5 * hits)
            lvl += step

        levels = [{'price': float(k), 'score': float(v)} for k, v in pools.items()]
        levels.sort(key=lambda x: (-x['score'], x['price']))
        return levels

    def select_target_pool(self, df: pd.DataFrame, i: int, direction: str, entry_price: float, pools: List[Dict]):
        if not pools:
            return None
        atr_series = self._atr_series(df)
        atr = atr_series.iloc[i] if i < len(atr_series) else np.nan
        max_dist = 1.5 * atr if not np.isnan(atr) else None

        if direction == 'LONG':
            candidates = [p for p in pools if p['price'] > entry_price]
            candidates.sort(key=lambda p: (-p['score'], abs(p['price'] - entry_price)))
            if max_dist is not None:
                within = [p for p in candidates if (p['price'] - entry_price) <= max_dist]
                if within:
                    return within[0]['price']
            return candidates[0]['price'] if candidates else None
        else:
            candidates = [p for p in pools if p['price'] < entry_price]
            candidates.sort(key=lambda p: (-p['score'], abs(p['price'] - entry_price)))
            if max_dist is not None:
                within = [p for p in candidates if (entry_price - p['price']) <= max_dist]
                if within:
                    return within[0]['price']
            return candidates[0]['price'] if candidates else None

    def check_long_setup(self, symbol: str, i: int) -> bool:
        df = self.dfs[symbol]
        if df is None or 'min' not in df.columns or 'fvg_bull_high' not in df.columns: return False # Asegurar que las columnas existan

        # 'i' es el índice de la vela actual (en desarrollo)
        # Buscamos estructura en velas cerradas (hasta i, excluyendo i)
        recent_lows = df['min'].iloc[i-50:i].dropna()
        if len(recent_lows) < 2 or recent_lows.iloc[-1] >= recent_lows.iloc[-2]: return False
        
        # sweep_idx es el índice numérico (iloc)
        try:
            sweep_idx = df.index.get_indexer([recent_lows.index[-1]])[0]
        except IndexError:
             return False # No se encontró el índice

        if i - sweep_idx > 12: return False
        
        fvg_window = df.iloc[sweep_idx:i] # FVGs en velas cerradas
        bullish_fvgs = fvg_window[['fvg_bull_low', 'fvg_bull_high']].dropna()
        if bullish_fvgs.empty: return False
        
        last_fvg = bullish_fvgs.iloc[-1]
        
        # Compara el 'low' de la vela actual (en desarrollo)
        if df['low'].iloc[i] <= last_fvg['fvg_bull_high']:
            # Filtro MACD (usar vela cerrada previa, i-1)
            if self.enable_macd_filter:
                if 'macd_hist' not in df.columns or i-2 < 0: return False
                mh_prev = df['macd_hist'].iloc[i-1]; m_prev = df['macd'].iloc[i-1]
                ms_prev = df['macd_signal'].iloc[i-1]; mh_prev2 = df['macd_hist'].iloc[i-2]
                if pd.isna(mh_prev) or pd.isna(m_prev) or pd.isna(ms_prev) or pd.isna(mh_prev2): return False
                if not (mh_prev > 0 and m_prev > ms_prev and mh_prev >= mh_prev2):
                    return False # Filtro MACD alcista no cumplido
            
            entry_price = float(last_fvg['fvg_bull_high'])
            liquidity_level = float(recent_lows.iloc[-1])
            # Pools 48h y selección de objetivo por concentración
            pools = self.build_liquidity_pools(df, i, lookback=self.pool_lookback_bars, tol=self.equal_tol)
            tp_pool = self.select_target_pool(df, i, 'LONG', entry_price, pools)
            
            logger.info(f"🔍 [{symbol}] Setup LONG detectado en vela {df.index[i]}. TP_pool={tp_pool}")
            asyncio.create_task(self.open_position(symbol, i, 'LONG', entry_price, liquidity_level, tp_override=tp_pool))
            return True
        return False

    def check_short_setup(self, symbol: str, i: int) -> bool:
        df = self.dfs[symbol]
        if df is None or 'max' not in df.columns or 'fvg_bear_low' not in df.columns: return False

        recent_highs = df['max'].iloc[i-50:i].dropna()
        if len(recent_highs) < 2 or recent_highs.iloc[-1] <= recent_highs.iloc[-2]: return False

        try:
            sweep_idx = df.index.get_indexer([recent_highs.index[-1]])[0]
        except IndexError:
             return False # No se encontró el índice

        if i - sweep_idx > 12: return False

        fvg_window = df.iloc[sweep_idx:i] # FVGs en velas cerradas
        bearish_fvgs = fvg_window[['fvg_bear_low', 'fvg_bear_high']].dropna()
        if bearish_fvgs.empty: return False
            
        last_fvg = bearish_fvgs.iloc[-1]
        
        # Compara el 'high' de la vela actual (en desarrollo)
        if df['high'].iloc[i] >= last_fvg['fvg_bear_low']:
            # Filtro MACD (usar vela cerrada previa, i-1)
            if self.enable_macd_filter:
                if 'macd_hist' not in df.columns or i-2 < 0: return False
                mh_prev = df['macd_hist'].iloc[i-1]; m_prev = df['macd'].iloc[i-1]
                ms_prev = df['macd_signal'].iloc[i-1]; mh_prev2 = df['macd_hist'].iloc[i-2]
                if pd.isna(mh_prev) or pd.isna(m_prev) or pd.isna(ms_prev) or pd.isna(mh_prev2): return False
                if not (mh_prev < 0 and m_prev < ms_prev and mh_prev <= mh_prev2):
                    return False # Filtro MACD bajista no cumplido

            entry_price = float(last_fvg['fvg_bear_low'])
            liquidity_level = float(recent_highs.iloc[-1])
            pools = self.build_liquidity_pools(df, i, lookback=self.pool_lookback_bars, tol=self.equal_tol)
            tp_pool = self.select_target_pool(df, i, 'SHORT', entry_price, pools)
            
            logger.info(f"🔍 [{symbol}] Setup SHORT detectado en vela {df.index[i]}. TP_pool={tp_pool}")
            asyncio.create_task(self.open_position(symbol, i, 'SHORT', entry_price, liquidity_level, tp_override=tp_pool))
            return True
        return False

    # --- Gestión de Órdenes y Posiciones (sin cambios) ---
    async def open_position(self, symbol: str, entry_idx: int, direction: str, entry_price: float, 
                            liquidity_level: float, # Mantenido por compatibilidad, pero no crucial para copy
                            tp_override: Optional[float] = None, 
                            is_copy: bool = False): # <-- Flag añadido
        df = self.dfs[symbol]
        if df is None: return
        
        # Obtener configuración específica del símbolo (si existe)
        symbol_config = self.symbol_configs.get(symbol, {})
        tp_percentage = symbol_config.get('tp_percentage', 2.0) 
        sl_percentage = symbol_config.get('sl_percentage', 1.0) 
        # risk_reward = symbol_config.get('risk_reward_ratio', self.risk_reward_ratio) # Ya no se usa directamente aquí

        # Calcular capital a arriesgar (esto es igual para SMC y Copy)
        capital_to_risk = self.balance * self.risk_per_trade_pct
        if self.balance <= 0 or capital_to_risk <= 0:
            logger.error(f"[{symbol}] Balance insuficiente (${self.balance:,.2f}) o riesgo configurado a cero. No se puede abrir posición.")
            await self.send_telegram_message(f"🚨 <b>Balance Insuficiente</b>\n\nNo se puede abrir {symbol} {direction}.\nBalance: ${self.balance:,.2f}")
            return

        # Inicializar variables
        stop_loss_price = 0.0
        take_profit_price = 0.0
        original_stop_loss_price = 0.0
        risk_per_unit = 0.0 
        position_size = 0.0

        if not is_copy:
            # --- Lógica para trades SMC (con SL/TP) ---
            log_prefix = "SMC"
            logger.info(f"[{symbol}] Calculando parámetros para trade {log_prefix}...")
            # Calcular SL y TP usando porcentajes (SOLO PARA TRADES SMC)
            if direction == 'LONG':
                stop_loss_price = entry_price * (1 - sl_percentage / 100)
                risk_per_unit = entry_price - stop_loss_price
                default_tp = entry_price * (1 + tp_percentage / 100) # TP basado en %
            else: # SHORT
                stop_loss_price = entry_price * (1 + sl_percentage / 100)
                risk_per_unit = stop_loss_price - entry_price
                default_tp = entry_price * (1 - tp_percentage / 100) # TP basado en %

            original_stop_loss_price = stop_loss_price # Guardamos el SL original

            if risk_per_unit <= 0:
                logger.warning(f"[{symbol}] Ignorando señal SMC, riesgo por unidad inválido ({risk_per_unit:.4f}) con SL a {stop_loss_price:.4f}. Señal en vela {df.index[entry_idx]}")
                return

            # Lógica de TP (pool o fijo por %)
            take_profit_price = default_tp # Empezar con el TP por defecto (%)
            if tp_override is not None:
                rr = 0.0
                # Asegurar que risk_per_unit > 0 para evitar división por cero
                if risk_per_unit > 0: 
                    if direction == 'LONG':
                        if tp_override > entry_price: 
                            rr = (tp_override - entry_price) / risk_per_unit
                    else: # SHORT
                         if tp_override < entry_price: 
                            rr = (entry_price - tp_override) / risk_per_unit
                
                # Usar TP del pool SOLO si cumple el R:R mínimo
                if rr >= self.min_rr:
                    take_profit_price = tp_override
                    logger.info(f"[{symbol}] Usando TP de Pool de Liquidez: ${take_profit_price:,.4f} (R:R {rr:.2f}x >= {self.min_rr:.2f}x)")
                else:
                    logger.info(f"[{symbol}] TP de Pool (${tp_override:,.4f}, R:R {rr:.2f}x < {self.min_rr:.2f}x) ignorado. Usando TP por % ({tp_percentage}%): ${take_profit_price:,.4f}")
            else:
                 logger.info(f"[{symbol}] No hay TP de Pool. Usando TP por % ({tp_percentage}%): ${take_profit_price:,.4f}")

            # Calcular tamaño para SMC basado en riesgo
            position_size = capital_to_risk / risk_per_unit

        else:
            # --- Lógica para trades COPY (sin SL/TP propios, tamaño diferente) ---
            log_prefix = "COPY"
            logger.info(f"[{symbol}] Calculando parámetros para trade {log_prefix}...")
            # Para copy trades, SL/TP son 0.0
            stop_loss_price = 0.0
            take_profit_price = 0.0
            original_stop_loss_price = 0.0
            
            # Calcular tamaño para COPY: 
            # Opción: Usar un % del balance como margen
            copy_margin_pct = self.risk_per_trade_pct # Reutilizamos el % riesgo como % de margen para copias
            lev = self.leverage_per_symbol.get(symbol, self.leverage) # Usar apalancamiento copiado o por defecto
            margin_to_use = self.balance * copy_margin_pct
            if entry_price > 0 and lev > 0:
                 position_size = (margin_to_use * lev) / entry_price
                 logger.info(f"[{symbol}] Tamaño para COPY trade: {position_size:.4f} (Margen: ${margin_to_use:.2f}, Lev: {lev}x)")
            else:
                 logger.error(f"[{symbol}] Precio de entrada ({entry_price}) o apalancamiento ({lev}) inválido para calcular tamaño de copy trade.")
                 return

            if position_size <= 0:
                logger.error(f"[{symbol}] Tamaño calculado para COPY trade es inválido ({position_size:.4f}). No se puede abrir.")
                return

        # --- Cálculos Comunes (Margen, Liquidación) ---
        lev = self.leverage_per_symbol.get(symbol, self.leverage) # Apalancamiento final
        margin_used = (position_size * entry_price) / lev if lev > 0 else (position_size * entry_price) # Evitar división por cero
        
        liquidation_price = 0.0
        if lev > 0:
            # Estimación simple de liquidación (puede variar según exchange)
            liquidation_pct = (1 / lev) * 0.95 # Factor de seguridad del 95% del margen
            liquidation_price = entry_price * (1 - liquidation_pct) if direction == 'LONG' else entry_price * (1 + liquidation_pct)
        else:
            logger.warning(f"[{symbol}] Apalancamiento es 0, no se puede calcular precio de liquidación.")

        # --- Crear y Guardar Objeto Position ---
        try:
             entry_time_dt = df.index[entry_idx].to_pydatetime()
             # Asegurarse que tiene timezone (UTC por defecto de ccxt)
             if entry_time_dt.tzinfo is None:
                  entry_time_dt = entry_time_dt.replace(tzinfo=pd.Timestamp(0, tz='UTC').tzinfo)

             pos = Position(
                symbol=symbol,
                direction=direction, 
                size=position_size, 
                entry_price=entry_price,
                entry_time=entry_time_dt, 
                entry_idx=entry_idx, # Guardar el índice original, aunque no se use mucho
                stop_loss=stop_loss_price,
                original_stop_loss=original_stop_loss_price, 
                take_profit=take_profit_price,
                liquidation_price=liquidation_price, 
                margin_used=margin_used,
                is_copy=is_copy # <-- Pasar el flag
            )
        except Exception as e:
            logger.error(f"Error creando objeto Position para {symbol}: {e}", exc_info=True)
            return

        # Guardar en memoria y persistencia
        self.positions[symbol] = pos
        # Guardar estado ANTES de enviar notificación/orden real
        await self._save_persistent_positions() 
        # Actualizar último tiempo de señal para evitar repetición en la misma vela
        self.last_signal_times[symbol] = df.index[entry_idx] 

        logger.info(f"📢 [{symbol}] Señal {log_prefix} {direction}: Tamaño {position_size:.4f} @ ${entry_price:.4f}")
        logger.info(f"   SL: ${stop_loss_price:.4f}, TP: ${take_profit_price:.4f}, Liq: ${liquidation_price:.4f}, Margen: ${margin_used:.2f}")

        # --- Lógica REAL para enviar orden ---
        # try:
        #    order_result = await self.client.place_order(
        #         symbol=symbol.replace('USDT','/USDT'), # Ajustar formato si es necesario
        #         side='buy' if direction == 'LONG' else 'sell', 
        #         order_type='market', # O 'limit' si prefieres
        #         quantity=position_size,
        #         # price=entry_price, # Solo para limit
        #         # stop_loss=stop_loss_price if not is_copy else None, # Enviar SL solo si no es copia
        #         # take_profit=take_profit_price if not is_copy else None # Enviar TP solo si no es copia
        #         # ... otros parámetros requeridos por BitunixClient ...
        #     ) 
        #    logger.info(f"[{symbol}] Orden {log_prefix} enviada a Bitunix: {order_result}")
        #    await self.notify_entry(pos) # Notificar SÓLO si la orden fue exitosa
        # except Exception as e:
        #    logger.error(f"[{symbol}] ¡¡ERROR AL ENVIAR ORDEN {log_prefix} A BITUNIX!!: {e}", exc_info=True)
        #    await self.send_telegram_message(f"🚨 <b>¡ERROR DE ORDEN!</b>\n\nNo se pudo abrir {symbol} {direction} en Bitunix.\nError: {e}")
        #    # Si falla la orden, revertir el estado (quitar de self.positions y guardar de nuevo)
        #    self.positions[symbol] = None
        #    await self._save_persistent_positions()
        #    return # No continuar si la orden falló
        # --- Fin Lógica REAL ---
            
        # --- Simulación (Eliminar o comentar para trading real) ---
        logger.warning(f"[{symbol}] SIMULACIÓN: Orden {log_prefix} {direction} NO enviada a Bitunix.")
        await self.notify_entry(pos) # Notificar (en simulación)
        # --- Fin Simulación ---
    async def manage_position(self, symbol: str, idx: int):
        """
        Gestiona la salida de un trade activo (SMC) usando la lógica HÍBRIDA (BASADA EN TIEMPO).
        IGNORA las posiciones abiertas por copy trading.
        """
        pos = self.positions.get(symbol) # Usar .get() para evitar KeyError si se eliminó
        
        # --- SALIDA TEMPRANA SI NO HAY POSICIÓN O ES COPIA ---
        if not pos or pos.is_copy:
            if pos and pos.is_copy:
                 logger.debug(f"[{symbol}] Ignorando gestión interna (es copia). Esperando señal de cierre de wallet.")
            # Si no hay posición (pos es None), no hacer nada
            return 
        # --- FIN SALIDA TEMPRANA ---

        # --- A partir de aquí, SÓLO se ejecuta para posiciones SMC ---
        
        df = self.dfs[symbol]
        if df is None: 
             logger.warning(f"[{symbol}] No hay DataFrame disponible para gestionar posición SMC.")
             return
        
        symbol_config = self.symbol_configs.get(symbol, {})
        enable_structural_stop = symbol_config.get('enable_structural_stop', True)
        
        # Validar índice antes de acceder
        if idx < 0 or idx >= len(df):
             logger.error(f"[{symbol}] Índice {idx} fuera de rango para DataFrame de tamaño {len(df)}. No se puede gestionar.")
             return
             
        candle = df.iloc[idx] # Vela actual en desarrollo
        current_timestamp = candle.name # Timestamp de la vela actual
        
        # Asegurar timezone en entry_time (importante si se cargó de JSON)
        if pos.entry_time.tzinfo is None:
            try:
                 # Asegurarse que current_timestamp es un Timestamp con tzinfo
                 if isinstance(current_timestamp, pd.Timestamp) and current_timestamp.tzinfo:
                      pos.entry_time = pos.entry_time.replace(tzinfo=current_timestamp.tzinfo)
                 else:
                      # Si no, usar UTC como fallback seguro
                      pos.entry_time = pos.entry_time.replace(tzinfo=pd.Timestamp(0, tz='UTC').tzinfo) 
                      logger.warning(f"[{symbol}] current_timestamp no tenía timezone. Usando UTC para pos.entry_time.")
            except AttributeError:
                 logger.error(f"[{symbol}] pos.entry_time ({pos.entry_time}) no es un objeto datetime válido. No se puede gestionar.")
                 return # Salir si el tiempo de entrada es inválido
            except Exception as e:
                 logger.error(f"[{symbol}] Error inesperado al asignar timezone a entry_time: {e}", exc_info=True)
                 return

        exit_reason, exit_price = None, 0
        new_sl_price = pos.stop_loss # Empezar con el SL actual (que NO es 0.0 para SMC)

        # --- Trailing Stop Estructural (Basado en Tiempo) - SOLO PARA SMC ---
        if enable_structural_stop and pos.stop_loss > 0: # Solo si hay un SL inicial definido
            try:
                # Filtrar DF desde entrada hasta vela ANTERIOR a la actual
                # Asegurarse que entry_time es menor o igual a current_timestamp
                if pos.entry_time <= current_timestamp:
                    df_since_entry = df.loc[pos.entry_time : current_timestamp].iloc[:-1]
                else:
                    logger.warning(f"[{symbol}] entry_time ({pos.entry_time}) es posterior a current_timestamp ({current_timestamp}). No se puede calcular trailing.")
                    df_since_entry = pd.DataFrame() # Dataframe vacío para evitar errores abajo
                
                # Asegurarse que el slice no esté vacío antes de buscar min/max
                if not df_since_entry.empty:
                    if pos.direction == 'LONG':
                        recent_structure_lows = df_since_entry['min'].dropna()
                        if not recent_structure_lows.empty:
                            new_protective_stop = recent_structure_lows.iloc[-1]
                            # Mover SL solo si el nuevo mínimo es MÁS ALTO que el SL actual Y MÁS ALTO que el precio de entrada (para asegurar profit o BE)
                            if new_protective_stop > pos.stop_loss and new_protective_stop >= pos.entry_price: 
                                new_sl_price = new_protective_stop
                    else: # SHORT
                        recent_structure_highs = df_since_entry['max'].dropna()
                        if not recent_structure_highs.empty:
                            new_protective_stop = recent_structure_highs.iloc[-1]
                            # Mover SL solo si el nuevo máximo es MÁS BAJO que el SL actual Y MÁS BAJO que el precio de entrada
                            if new_protective_stop < pos.stop_loss and new_protective_stop <= pos.entry_price: 
                                new_sl_price = new_protective_stop
                # else: # No es necesario loguear si está vacío, es normal al principio
                #    logger.debug(f"[{symbol}] df_since_entry vacío, no se puede calcular trailing stop.")

            except KeyError as ke:
                 logger.warning(f"[{symbol}] Error de índice de tiempo al buscar trailing stop ({ke}). Puede ocurrir si entry_time no está en el índice. Slice: {pos.entry_time} -> {current_timestamp}")
            except Exception as e:
                 logger.error(f"[{symbol}] Error inesperado calculando trailing stop: {e}", exc_info=True)

            # Si el nuevo SL calculado es diferente Y válido (mayor que 0), actualizar y notificar
            if new_sl_price > 0 and new_sl_price != pos.stop_loss:
                old_sl = pos.stop_loss
                pos.stop_loss = new_sl_price # Actualiza el SL dinámico en memoria
                logger.info(f"[{symbol}] 🛡️ TRAILING STOP ESTRUCTURAL (SMC): SL movido de ${old_sl:,.4f} a ${new_sl_price:,.4f}")
                
                # --- Lógica REAL para modificar orden SL en el exchange ---
                # try:
                #    # Asegurarse de tener el ID de la orden SL original para modificarla
                #    # sl_order_id = ... # Necesitas almacenar el ID al crear la orden SL
                #    # await self.client.modify_sl_order(symbol, sl_order_id, new_sl_price, ...) 
                #    logger.info(f"[{symbol}] Orden SL modificada en Bitunix a ${new_sl_price:,.4f}")
                #    msg = (
                #        f"🛡️ <b>TRAILING STOP ACTUALIZADO ({symbol} - SMC)</b>\n\n"
                #        f"Posición: <b>{pos.direction}</b>\n"
                #        f"El Stop Loss se ha movido de ${old_sl:,.4f} a <b>${new_sl_price:,.4f}</b>"
                #    )
                #    asyncio.create_task(self.send_telegram_message(msg))
                #    # Guardar el estado actualizado SOLO si la modificación fue exitosa
                #    await self._save_persistent_positions() 
                # except Exception as e:
                #    logger.error(f"[{symbol}] ¡¡ERROR AL MODIFICAR SL EN BITUNIX!!: {e}", exc_info=True)
                #    await self.send_telegram_message(f"🚨 <b>¡ERROR AL MODIFICAR SL!</b>\n\nNo se pudo mover el SL de {symbol} a ${new_sl_price:,.4f}.\nError: {e}")
                #    # Revertir el cambio en memoria si falló la modificación en el exchange
                #    pos.stop_loss = old_sl 
                # --- Fin Lógica REAL ---

                # --- Simulación ---
                logger.warning(f"[{symbol}] SIMULACIÓN: Orden SL NO modificada en Bitunix.")
                msg = (
                     f"🛡️ <b>TRAILING STOP ACTUALIZADO ({symbol} - SMC)</b>\n\n"
                     f"Posición: <b>{pos.direction}</b>\n"
                     f"El Stop Loss se ha movido de ${old_sl:,.4f} a <b>${new_sl_price:,.4f}</b>"
                 )
                asyncio.create_task(self.send_telegram_message(msg))
                # Guardar el estado actualizado (en simulación, siempre guardar)
                await self._save_persistent_positions()
                # --- Fin Simulación ---


        # --- Lógica de Salida (SL, TP, Time Limit) - SOLO PARA SMC ---
        # Solo comprobar si hay un SL/TP válido definido (mayor que 0)
        # Usamos el pos.stop_loss actualizado (podría ser el trailing)
        if pos.direction == 'LONG':
            # TP Check (solo si TP > 0)
            if not exit_reason and pos.take_profit > 0 and candle['high'] >= pos.take_profit: 
                exit_reason, exit_price = 'Take Profit', pos.take_profit
            # SL Check (solo si SL > 0)
            elif not exit_reason and pos.stop_loss > 0 and candle['low'] <= pos.stop_loss: 
                # Determinar si es SL original o estructural
                exit_reason = 'Stop Loss' if abs(pos.stop_loss - pos.original_stop_loss) < 1e-9 else 'Stop Estructural' 
                exit_price = pos.stop_loss
        
        else: # SHORT
            # TP Check (solo si TP > 0)
            if not exit_reason and pos.take_profit > 0 and candle['low'] <= pos.take_profit: 
                exit_reason, exit_price = 'Take Profit', pos.take_profit
            # SL Check (solo si SL > 0)
            elif not exit_reason and pos.stop_loss > 0 and candle['high'] >= pos.stop_loss: 
                exit_reason = 'Stop Loss' if abs(pos.stop_loss - pos.original_stop_loss) < 1e-9 else 'Stop Estructural'
                exit_price = pos.stop_loss
        
        # Comprobación de tiempo (Basado en conteo por Tiempo) - SOLO PARA SMC
        if not exit_reason and self.max_candles_in_trade > 0:
            try:
                 # Contar velas desde la entrada HASTA la vela actual INCLUSIVE
                 if pos.entry_time <= current_timestamp:
                      candles_in_trade = len(df.loc[pos.entry_time : current_timestamp]) 
                      # La condición es >= max_candles + 1 porque cuenta la vela de entrada
                      # Si max=24, debe cerrar DESPUÉS de completarse la vela 24 (índice 24), o sea, en la vela 25 (índice 25)
                      if candles_in_trade >= self.max_candles_in_trade + 1: 
                           exit_reason, exit_price = 'Time Limit', candle['close']
                           logger.info(f"[{symbol}] Límite de tiempo alcanzado ({candles_in_trade-1} velas completas >= {self.max_candles_in_trade}). Cerrando al cierre de vela actual.")
                 else:
                      logger.warning(f"[{symbol}] entry_time posterior a current_timestamp, no se puede comprobar límite de tiempo.")
            except KeyError as ke:
                 logger.warning(f"[{symbol}] Error de índice de tiempo al comprobar límite de velas ({ke}). Slice: {pos.entry_time} -> {current_timestamp}")
            except Exception as e:
                 logger.error(f"[{symbol}] Error inesperado comprobando límite de tiempo: {e}", exc_info=True)


        # Cierre - SOLO PARA SMC 
        if exit_reason:
            logger.info(f"[{symbol}] Condición de cierre SMC detectada: {exit_reason} @ ${exit_price:.4f}")
            # --- Lógica REAL para CERRAR A MERCADO en Bitunix ---
            # try:
            #    # Asegurarse de cancelar órdenes SL/TP pendientes antes de cerrar a mercado
            #    # await self.client.cancel_all_orders(symbol) o similar
            #    close_result = await self.client.close_position(symbol, ...) # O crear orden market opuesta
            #    logger.info(f"[{symbol}] Orden de cierre SMC enviada a Bitunix: {close_result}")
            #    # Solo actualizar estado interno si el cierre en exchange fue exitoso
            #    await self.close_position(symbol, exit_price, reason=f"SMC {exit_reason}") 
            # except Exception as e:
            #    logger.error(f"[{symbol}] ¡¡ERROR AL CERRAR POSICIÓN SMC EN BITUNIX!!: {e}", exc_info=True)
            #    await self.send_telegram_message(f"🚨 <b>¡ERROR AL CERRAR!</b>\n\nNo se pudo cerrar {symbol} en Bitunix por {exit_reason}.\nError: {e}")
            #    # NO llamar a self.close_position si falla, para reintentar en el próximo ciclo.
            # --- Fin Lógica REAL ---

            # --- Simulación (Eliminar o comentar para trading real) ---
            logger.warning(f"[{symbol}] SIMULACIÓN: Orden de cierre SMC ({exit_reason}) NO enviada a Bitunix.")
            await self.close_position(symbol, exit_price, reason=f"SMC {exit_reason}")
            # --- Fin Simulación ---
    
    # <<< FUNCIÓN MODIFICADA >>>
    async def run(self):
        await self.send_telegram_message(
            f"🚀 <b>Bot SMC (Multi-Símbolo) Iniciado</b>\n\n"
            f"📉 Operando en: {', '.join(self.symbols)}\n"
            f"💼 Balance Inicial: ${self.balance:,.2f}\n\n"
            f"💬 Comandos disponibles:\n"
            f"   • /wallet - Estado de la wallet rastreada"
        )
        self.is_running = True
        
        # Iniciar listener de comandos de Telegram en background
        asyncio.create_task(self.listen_telegram_commands())

        while self.is_running:
            try:
                for symbol in self.symbols:
                    try:
                        if not await self.update_market_data(symbol):
                            logger.warning(f"[{symbol}] No se pudieron actualizar datos, saltando ciclo.")
                            continue 

                        df = self.dfs[symbol]
                        if df is None: continue
                        
                        # --- INICIO MODIFICACIÓN ---
                        # Ahora 'idx' es la vela actual, en desarrollo
                        idx = len(df) - 1 
                        # --- FIN MODIFICACIÓN ---
                        
                        if idx < self.structure_lookback + 2:
                            logger.warning(f"[{symbol}] Esperando más datos históricos...")
                            continue

                        current_candle_time = df.index[idx]
                        current_price = df['close'].iloc[idx] # El precio de cierre de la vela actual es el precio 'live'
                        
                        logger.info(f"🕵️ [{symbol}] Analizando vela {current_candle_time.strftime('%H:%M:%S')} (Precio actual: ${current_price:,.4f})...")

                        if self.positions[symbol]:
                            # La gestión de SL/TP usa el 'idx' actual (en desarrollo)
                            await self.manage_position(symbol, idx)
                        else:
                            # Límite global de posiciones concurrentes
                            open_count = sum(1 for p in self.positions.values() if p is not None)
                            if open_count >= self.max_concurrent_open:
                                logger.info(f"↩️ Límite de {self.max_concurrent_open} posiciones alcanzado, saltando nuevas entradas.")
                                continue
                            # --- INICIO MODIFICACIÓN (Control de vela repetida) ---
                            # Usamos el time de la vela actual para el control
                            last_time = self.last_signal_times[symbol]
                            
                            # Si ya procesamos una señal en esta misma vela en desarrollo
                            if last_time and last_time == current_candle_time:
                                logger.info(f"   [{symbol}] Setup de la vela {last_time} ya procesado, esperando siguiente vela...")
                                continue
                            # --- FIN MODIFICACIÓN ---

                            # El chequeo de setup usa el 'idx' actual (en desarrollo)
                            if self.check_long_setup(symbol, idx):
                                pass
                            elif self.check_short_setup(symbol, idx):
                                pass
                            else:
                                logger.info(f"   [{symbol}] No se encontraron setups válidos en la vela actual.")
                    
                    except Exception as e:
                        logger.error(f"Error procesando el símbolo {symbol}: {e}", exc_info=True)
                        await self.send_telegram_message(f"🚨 <b>Error Crítico en {symbol}</b>\n{e}")
                
                logger.debug(f"Ciclo completado. Esperando {self.refresh_seconds} segundos...")
                await asyncio.sleep(self.refresh_seconds)

            except Exception as e:
                logger.error(f"Error en el bucle principal (externo): {e}", exc_info=True)
                await self.send_telegram_message(f"🚨 <b>Error Crítico en el Bot (Bucle Principal)</b>\n\n{e}\n\nReintentando en 60 segundos.")
                await asyncio.sleep(60)

    def stop(self):
        self.is_running = False
        if self.wallet_tracker:
            asyncio.create_task(self.wallet_tracker.stop_monitoring())
    
    async def start_wallet_tracking(self, wallet_address: str, etherscan_api_key: Optional[str] = None, moralis_api_key: Optional[str] = None, debank_api_key: Optional[str] = None, enable_copy_trading: bool = True):
        """Inicia el rastreo de una wallet específica."""
        try:
            # Cargar API keys desde advanced_wallet_config.json si existe
            api_keys = self.load_advanced_wallet_config()
            
            # Usar API keys del archivo de configuración si están disponibles
            final_etherscan_key = api_keys.get('etherscan_api_key') or etherscan_api_key
            final_moralis_key = api_keys.get('moralis_api_key') or moralis_api_key
            final_debank_key = api_keys.get('debank_api_key') or debank_api_key
            
            logger.info(f"🔑 Usando API keys:")
            logger.info(f"   Etherscan: {'✅ Configurada' if final_etherscan_key else '❌ No disponible'}")
            logger.info(f"   Moralis: {'✅ Configurada' if final_moralis_key else '❌ No disponible'}")
            logger.info(f"   DeBank: {'✅ Configurada' if final_debank_key else '❌ No disponible'}")
            
            self.wallet_tracker = WalletTracker(
                wallet_address=wallet_address,
                telegram_bot=self.telegram_bot,
                telegram_chat_id=self.telegram_chat_id,
                etherscan_api_key=final_etherscan_key,
                moralis_api_key=final_moralis_key,
                debank_api_key=final_debank_key,
                check_interval=0.5  # Verificar cada 500ms para mínimo delay
            )
            
            # Habilitar copy trading si está solicitado
            if enable_copy_trading:
                self.wallet_tracker.enable_copy_trading(
                    open_callback=self.copy_open_position,
                    close_callback=self.copy_close_position
                )
            
            # Iniciar monitoreo en background
            asyncio.create_task(self.wallet_tracker.start_monitoring())
            
            logger.info(f"🔍 Rastreo de wallet iniciado: {wallet_address}")
            
            # Log detallado de inicio
            logger.info("📊 ESTADO ACTUAL DEL BOT SMC:")
            
            # Obtener información de posiciones abiertas actuales del BOT
            bot_open_positions = []
            for symbol, position in self.positions.items():
                if position is not None:
                    bot_open_positions.append({
                        'symbol': symbol,
                        'direction': position.direction,
                        'entry_price': position.entry_price,
                        'size': position.size,
                        'leverage': self.leverage_per_symbol.get(symbol, 15)
                    })
            
            # Log de posiciones del bot
            logger.info(f"   💼 Balance bot: ${self.balance:.2f}")
            logger.info(f"   📈 Posiciones bot abiertas: {len(bot_open_positions)}")
            
            if bot_open_positions:
                for i, pos in enumerate(bot_open_positions, 1):
                    logger.info(f"   📊 Bot Posición {i}: {pos['symbol']} | {pos['direction']} | "
                              f"${pos['entry_price']:.2f} | {pos['leverage']}x | Tamaño: {pos['size']:.4f}")
            else:
                logger.info("   ✅ Bot sin posiciones activas")
            
            logger.info(f"   🎯 Símbolos monitoreados: {', '.join(self.symbols)}")
            logger.info(f"   🔄 Copy trading: {'HABILITADO' if enable_copy_trading else 'DESHABILITADO'}")
            
            # Obtener posiciones de la WALLET rastreada
            logger.info("=" * 60)
            logger.info("👁️ CONSULTANDO POSICIONES DE LA WALLET RASTREADA:")
            
            try:
                # Consultar posiciones actuales de la wallet
                wallet_positions = await self.get_wallet_current_positions(wallet_address)
                
                if wallet_positions:
                    logger.info(f"   📊 Posiciones wallet activas: {len(wallet_positions)}")
                    for i, pos in enumerate(wallet_positions, 1):
                        logger.info(f"   🎯 Wallet Posición {i}: {pos['symbol']} | {pos['direction']} | "
                                  f"${pos['entry_price']:.2f} | {pos['leverage']}x | Tamaño: {pos['size']:.4f}")
                else:
                    logger.info("   ✅ Wallet sin posiciones activas detectadas")
            except Exception as e:
                logger.warning(f"   ⚠️ No se pudieron obtener posiciones de la wallet: {e}")
            
            logger.info("=" * 60)
            
            # Crear mensaje con posiciones abiertas (WALLET y BOT)
            copy_status = "✅ HABILITADO" if enable_copy_trading else "❌ DESHABILITADO"
            
            message = (
                f"🔍 <b>RASTREO DE WALLET ACTIVADO</b>\n\n"
                f"📍 Wallet: <code>{wallet_address}</code>\n"
                f"🔄 Copy Trading: {copy_status}\n"
                f"⏱️ Intervalo: 500ms (ultra-rápido)\n\n"
            )
            
            # Agregar posiciones de la WALLET rastreada
            try:
                wallet_positions = await self.get_wallet_current_positions(wallet_address)
                message += f"👁️ <b>POSICIONES DE LA WALLET RASTREADA:</b>\n"
                
                if wallet_positions:
                    message += f"Cantidad posiciones wallet: {len(wallet_positions)}\n\n"
                    for i, pos in enumerate(wallet_positions, 1):
                        message += (
                            f"<b>Wallet Posición {i}:</b>\n"
                            f"Símbolo: {pos['symbol']}\n"
                            f"Dirección: {pos['direction']}\n"
                            f"Precio entrada: ${pos['entry_price']:.2f}\n"
                            f"Tamaño: {pos['size']:.4f}\n"
                            f"Apalancamiento: {pos['leverage']}x\n"
                            f"Plataforma: {pos['platform']}\n\n"
                        )
                else:
                    message += "Cantidad posiciones wallet: 0\n"
                    message += "Wallet sin posiciones activas\n\n"
            except:
                message += "⚠️ No se pudieron consultar posiciones de la wallet\n\n"
            
            # Agregar posiciones del BOT SMC
            message += f"🤖 <b>POSICIONES DEL BOT SMC:</b>\n"
            if bot_open_positions:
                message += f"Cantidad posiciones bot: {len(bot_open_positions)}\n\n"
                for i, pos in enumerate(bot_open_positions, 1):
                    message += (
                        f"<b>Bot Posición {i}:</b>\n"
                        f"Símbolo: {pos['symbol']}\n"
                        f"Dirección: {pos['direction']}\n"
                        f"Precio entrada: ${pos['entry_price']:.2f}\n"
                        f"Tamaño: {pos['size']:.4f}\n"
                        f"Apalancamiento: {pos['leverage']}x\n\n"
                    )
            else:
                message += "Cantidad posiciones bot: 0\n"
                message += "Bot sin posiciones activas\n\n"
            
            message += "🎯 Monitoreando wallet para copy trading..."
            
            await self.send_telegram_message(message)
            
        except Exception as e:
            logger.error(f"Error iniciando rastreo de wallet: {e}")
            await self.send_telegram_message(f"🚨 Error iniciando rastreo de wallet: {e}")
    
    async def copy_open_position(self, tracked_position):
        """Callback para abrir posición copiando la wallet rastreada."""
        try:
            original_symbol = tracked_position.symbol.upper()
            
            # Intentar diferentes formatos de símbolo
            possible_symbols = [
                f"{original_symbol}USDT", f"{original_symbol}/USDT", 
                original_symbol, f"{original_symbol}USD"
            ]
            
            # Buscar el símbolo en nuestra lista de símbolos *actualmente* monitoreados
            symbol = None
            current_monitored_symbols = list(self.symbols) # Copia para iterar seguro
            for possible in possible_symbols:
                # Comprobación más flexible (ignora '/')
                normalized_possible = possible.replace('/', '')
                normalized_monitored = [s.replace('/', '') for s in current_monitored_symbols]
                if normalized_possible in normalized_monitored:
                    # Encontrar el símbolo original con el formato correcto
                    idx = normalized_monitored.index(normalized_possible)
                    symbol = current_monitored_symbols[idx]
                    break
            
            # Si no está en nuestra lista monitoreada, ignorar
            if not symbol:
                 logger.warning(f"Símbolo {original_symbol} (buscado como {possible_symbols}) de wallet rastreada no está en la lista de símbolos monitoreados: {current_monitored_symbols}. Ignorando copy trade.")
                 return

            # Verificar si ya tenemos posición en este símbolo
            # Usamos `is not None` para asegurarnos
            if self.positions.get(symbol) is not None:
                pos_info = self.positions[symbol]
                logger.warning(f"Ya tenemos posición abierta en {symbol} ({pos_info.direction}, Copy={pos_info.is_copy}, Entry=${pos_info.entry_price:.4f}). Ignorando nuevo copy trade.")
                return
            
            # Verificar límite de posiciones concurrentes
            open_count = sum(1 for p in self.positions.values() if p is not None)
            if open_count >= self.max_concurrent_open:
                logger.warning(f"Límite de {self.max_concurrent_open} posiciones concurrentes alcanzado ({open_count}). Ignorando copy trade para {symbol}.")
                await self.send_telegram_message(f"⚠️ Límite de {self.max_concurrent_open} posiciones alcanzado. No se pudo copiar {symbol} {tracked_position.position_type}.")
                return
            
            # Obtener datos actuales del mercado para el símbolo detectado
            if not await self.update_market_data(symbol):
                logger.error(f"No se pudieron obtener datos de mercado para {symbol}. No se puede ejecutar copy trade.")
                return
            
            df = self.dfs[symbol]
            if df is None or len(df) < 2:
                logger.error(f"Datos de mercado insuficientes para {symbol} después de actualizar. No se puede ejecutar copy trade.")
                return
            
            # Usar precio actual (último cierre disponible)
            current_price = df['close'].iloc[-1]
            direction = tracked_position.position_type.upper() # Asegurar que sea 'LONG' o 'SHORT'
            
            # USAR EL MISMO APALANCAMIENTO que la wallet rastreada si está disponible
            wallet_leverage = tracked_position.leverage
            final_leverage = self.leverage_per_symbol.get(symbol, self.leverage) # Empezar con default
            if wallet_leverage is not None and wallet_leverage > 0:
                final_leverage = int(wallet_leverage)
                # Actualizar temporalmente el apalancamiento para este símbolo ANTES de abrir
                self.leverage_per_symbol[symbol] = final_leverage 
                logger.info(f"🔄 Usando apalancamiento de la wallet: {final_leverage}x para {symbol}")
            else:
                logger.info(f"ℹ️ Usando apalancamiento configurado: {final_leverage}x para {symbol} (no detectado en wallet).")

            logger.info(f"🔄 EJECUTANDO COPY TRADE: Abriendo {direction} en {symbol} @ ${current_price:.4f} con {final_leverage}x leverage...")
            
            # Llamar a open_position marcándola como copia
            # Pasar 0.0 para liquidity_level y None para tp_override ya que no se usan en copy
            await self.open_position(symbol=symbol, 
                                     entry_idx=len(df)-1, 
                                     direction=direction, 
                                     entry_price=current_price, 
                                     liquidity_level=0.0,    
                                     tp_override=None,       
                                     is_copy=True)          
            
            # Log adicional si la apertura fue procesada (el mensaje a TG ya está en open_position)
            # Verificar si la posición realmente se abrió (puede fallar por balance, etc.)
            if self.positions.get(symbol) and self.positions[symbol].is_copy:
                 logger.info(f"Apertura de Copy trade para {symbol} {direction} procesada correctamente.")
                 # (No enviamos otro mensaje a TG aquí para evitar duplicados)
            else:
                 logger.warning(f"Algo falló después de intentar abrir copy trade para {symbol}. Verificar logs anteriores.")


        except Exception as e:
            # Captura errores más generales que puedan ocurrir en el callback
            symbol_name = symbol if 'symbol' in locals() else original_symbol
            logger.error(f"Error CRÍTICO en callback copy_open_position para {symbol_name}: {e}", exc_info=True)
            try: 
                 await self.send_telegram_message(f"🚨 Error crítico procesando apertura de copy trade para {symbol_name}: {e}")
            except Exception as tg_err:
                 logger.error(f"Fallo al notificar error de copy_open_position por Telegram: {tg_err}")
    async def copy_close_position(self, tracked_position):
        """Callback para cerrar posición copiando la wallet rastreada."""
        try:
            # Mapear símbolo automáticamente - CUALQUIER símbolo
            original_symbol = tracked_position.symbol.upper()
            
            # Buscar el símbolo en nuestras posiciones activas
            symbol = None
            for active_symbol in self.positions.keys():
                if (active_symbol.replace('USDT', '').replace('/USDT', '') == original_symbol or
                    active_symbol == original_symbol or
                    active_symbol == f"{original_symbol}USDT"):
                    symbol = active_symbol
                    break
            
            # Si no encontramos el símbolo exacto, usar el formato estándar
            if not symbol:
                symbol = f"{original_symbol}USDT"
            
            # Verificar si tenemos posición abierta en este símbolo
            if not self.positions.get(symbol):
                logger.warning(f"No tenemos posición abierta en {symbol} para cerrar")
                return
            
            # Obtener precio actual
            if not await self.update_market_data(symbol):
                logger.error(f"No se pudieron obtener datos de {symbol} para copy trading")
                return
            
            df = self.dfs[symbol]
            if df is None or len(df) < 1:
                logger.error(f"Datos insuficientes para {symbol}")
                return
            
            current_price = df['close'].iloc[-1]
            
            logger.info(f"🔄 COPY TRADING: Cerrando posición en {symbol} @ ${current_price:.4f}")
            
            # Cerrar posición
            await self.close_position(symbol, current_price, "Copy Trading Close")
            
            await self.send_telegram_message(
                f"🔄 <b>COPY TRADING - POSICIÓN CERRADA</b>\n\n"
                f"📊 Símbolo: {symbol}\n"
                f"💰 Precio cierre: ${current_price:.4f}\n"
                f"⚡ Siguiendo wallet rastreada"
            )
            
        except Exception as e:
            logger.error(f"Error en copy trading (cerrar): {e}")
            await self.send_telegram_message(f"🚨 Error en copy trading (cerrar): {e}")
    
    async def handle_wallet_command(self, message_id=None):
        """Maneja el comando /wallet para mostrar estado actual de la wallet rastreada."""
        if not self.wallet_tracker:
            await self.send_telegram_message(
                "⚠️ <b>Rastreo de Wallet No Activo</b>\n\n"
                "No hay ninguna wallet siendo rastreada en este momento."
            )
            return
        
        try:
            # Obtener posiciones actuales desde Hyperliquid
            positions = await self.wallet_tracker.get_current_positions()
            
            if not positions:
                await self.send_telegram_message(
                    f"📊 <b>ESTADO DE LA WALLET</b>\n\n"
                    f"📍 Wallet: <code>{self.wallet_tracker.wallet_address}</code>\n"
                    f"💼 Posiciones Activas: 0\n\n"
                    f"✅ Wallet sin posiciones abiertas"
                )
                return
            
            # Construir mensaje con todas las posiciones
            message = (
                f"📊 <b>ESTADO DE LA WALLET</b>\n\n"
                f"📍 Wallet: <code>{self.wallet_tracker.wallet_address}</code>\n"
                f"💼 Posiciones Activas: {len(positions)}\n"
                f"⏰ Última actualización: {datetime.now().strftime('%H:%M:%S')}\n\n"
            )
            
            # Consultar datos adicionales de Hyperliquid para PnL
            import aiohttp
            url = "https://api.hyperliquid.xyz/info"
            headers = {"Content-Type": "application/json"}
            user_payload = {"type": "clearinghouseState", "user": self.wallet_tracker.wallet_address}
            
            total_pnl = 0
            total_margin = 0
            
            async with aiohttp.ClientSession() as session:
                async with session.post(url, headers=headers, json=user_payload) as response:
                    if response.status == 200:
                        data = await response.json()
                        margin_summary = data.get("marginSummary", {})
                        total_margin = float(margin_summary.get("totalMargin", 0))
                        total_pnl = float(margin_summary.get("totalUnrealizedPnl", 0))
                        
                        # Agregar resumen de cuenta
                        message += (
                            f"💰 <b>Resumen de Cuenta:</b>\n"
                            f"   • Margen Total: ${total_margin:,.2f}\n"
                            f"   • PnL No Realizado: ${total_pnl:,.2f}\n\n"
                        )
                        
                        # Obtener detalles de cada posición
                        positions_data = data.get("assetPositions", [])
                        
                        for idx, item in enumerate(positions_data, 1):
                            position_data = item.get("position", {})
                            symbol = position_data.get("coin", "UNKNOWN")
                            size = float(position_data.get("szi", 0))
                            
                            if size == 0:
                                continue
                            
                            direction = "🟢 LONG" if size > 0 else "🔴 SHORT"
                            entry_price = float(position_data.get("entryPx", 0))
                            position_value = float(position_data.get("positionValue", 0))
                            unrealized_pnl = float(position_data.get("unrealizedPnl", 0))
                            liq_price = float(position_data.get("liquidationPx", 0))
                            
                            leverage_info = position_data.get("leverage", {})
                            leverage_type = leverage_info.get("type", "cross")
                            leverage_value = leverage_info.get("value", 1)
                            
                            leverage_str = f"Cross ({leverage_value}x)" if leverage_type == "cross" else f"{leverage_value}x"
                            
                            pnl_emoji = "🟢" if unrealized_pnl >= 0 else "🔴"
                            
                            message += (
                                f"<b>#{idx} {symbol}</b> {direction}\n"
                                f"   • Tamaño: {abs(size):.4f}\n"
                                f"   • Precio Entrada: ${entry_price:,.2f}\n"
                                f"   • Valor Posición: ${position_value:,.2f}\n"
                                f"   • Apalancamiento: {leverage_str}\n"
                                f"   • {pnl_emoji} PnL: ${unrealized_pnl:,.2f}\n"
                                f"   • Precio Liq.: ${liq_price:,.2f}\n\n"
                            )
            
            # Crear botón de refresh
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Refresh", callback_data="refresh_wallet")]
            ])
            
            # Si es una actualización (presionaron el botón), editar el mensaje
            if message_id:
                try:
                    await self.telegram_bot.edit_message_text(
                        chat_id=self.telegram_chat_id,
                        message_id=message_id,
                        text=message,
                        parse_mode='HTML',
                        reply_markup=keyboard
                    )
                except Exception as e:
                    logger.error(f"Error editando mensaje: {e}")
            else:
                # Si es un comando nuevo, enviar mensaje nuevo
                await self.telegram_bot.send_message(
                    chat_id=self.telegram_chat_id,
                    text=message,
                    parse_mode='HTML',
                    reply_markup=keyboard
                )
            
        except Exception as e:
            logger.error(f"Error obteniendo estado de wallet: {e}")
            await self.send_telegram_message(
                f"🚨 <b>Error al obtener estado de wallet</b>\n\n"
                f"Error: {str(e)}"
            )
    
    async def get_wallet_status(self) -> Dict:
        """Obtiene el estado actual del rastreador de wallet."""
        if not self.wallet_tracker:
            return {'status': 'inactive', 'message': 'Rastreo de wallet no iniciado'}
        
        try:
            summary = await self.wallet_tracker.get_wallet_summary()
            return {'status': 'active', 'summary': summary}
        except Exception as e:
            logger.error(f"Error obteniendo estado de wallet: {e}")
            return {'status': 'error', 'message': str(e)}
    
    async def get_wallet_current_positions(self, wallet_address: str):
        """Obtiene las posiciones actuales de la wallet rastreada consultando directamente las APIs."""
        try:
            wallet_positions = []
            
            # Método 1: Usar Moralis API con tu API key real
            logger.info("🔍 Método 1: Consultando Moralis API con API key real...")
            if hasattr(self, 'wallet_tracker') and self.wallet_tracker and self.wallet_tracker.moralis_api_key:
                try:
                    import aiohttp
                    async with aiohttp.ClientSession() as session:
                        headers = {'X-API-Key': self.wallet_tracker.moralis_api_key}
                        url = f"https://deep-index.moralis.io/api/v2.2/{wallet_address}/erc20"
                        params = {'chain': 'eth'}
                        
                        async with session.get(url, headers=headers, params=params) as response:
                            if response.status == 200:
                                data = await response.json()
                                positions = self.parse_moralis_balances(data)
                                wallet_positions.extend(positions)
                                logger.info(f"✅ Moralis API encontró {len(positions)} posiciones REALES")
                                for pos in positions:
                                    logger.info(f"   📊 {pos['symbol']}: {pos['size']:.4f} ({pos['platform']})")
                            else:
                                logger.info(f"❌ Moralis API error: {response.status}")
                except Exception as e:
                    logger.info(f"❌ Error consultando Moralis: {e}")
            else:
                logger.info("❌ Moralis API key no disponible")
            
            # Método 2: Consultar HyperDash como respaldo
            if not wallet_positions:
                logger.info("🔍 Método 2: Consultando HyperDash API...")
                try:
                    async with aiohttp.ClientSession() as session:
                        url = f"https://hyperdash.info/api/trader/{wallet_address}"
                        async with session.get(url) as response:
                            if response.status == 200:
                                data = await response.json()
                                
                                # Procesar posiciones de HyperDash
                                if 'positions' in data:
                                    for pos_data in data['positions']:
                                        if pos_data.get('size', 0) > 0:  # Solo posiciones activas
                                            wallet_positions.append({
                                                'symbol': pos_data.get('symbol', 'UNKNOWN'),
                                                'direction': pos_data.get('side', 'LONG').upper(),
                                                'entry_price': float(pos_data.get('entry_price', 0)),
                                                'size': float(pos_data.get('size', 0)),
                                                'leverage': float(pos_data.get('leverage', 1)),
                                                'platform': 'HyperDash'
                                            })
                                    logger.info(f"✅ HyperDash encontró {len(wallet_positions)} posiciones")
                                else:
                                    logger.info("❌ HyperDash no devolvió posiciones")
                            else:
                                logger.info(f"❌ HyperDash API error: {response.status}")
                except Exception as e:
                    logger.info(f"❌ Error consultando HyperDash: {e}")
            
            
            # Método 3: Usar API alternativas más confiables
            if not wallet_positions:
                logger.info("Intentando APIs alternativas más confiables...")
                wallet_positions = await self.detect_positions_alternative_apis(wallet_address)
            
            # Método 4: Intentar detectar posiciones usando análisis de transacciones recientes
            if not wallet_positions:
                logger.info("Intentando detectar posiciones mediante análisis de transacciones...")
                wallet_positions = await self.detect_positions_from_transactions(wallet_address)
            
            # Método 5: Usar datos del wallet tracker si está disponible
            if not wallet_positions and hasattr(self, 'wallet_tracker') and self.wallet_tracker:
                try:
                    tracker_summary = await self.wallet_tracker.get_wallet_summary()
                    if tracker_summary.get('positions'):
                        for pos in tracker_summary['positions']:
                            wallet_positions.append({
                                'symbol': pos.symbol,
                                'direction': pos.position_type,
                                'entry_price': pos.entry_price,
                                'size': pos.size,
                                'leverage': pos.leverage,
                                'platform': pos.platform
                            })
                        logger.info(f"Posiciones obtenidas del wallet tracker: {len(wallet_positions)}")
                except Exception as e:
                    logger.debug(f"Error obteniendo posiciones del tracker: {e}")
            
            return wallet_positions
            
        except Exception as e:
            logger.error(f"Error obteniendo posiciones de wallet: {e}")
            # En caso de error, devolver lista vacía en lugar de datos hardcodeados
            return []
    
    def parse_moralis_balances(self, moralis_data):
        """Parsea datos de Moralis para extraer posiciones de tokens."""
        try:
            positions = []
            
            for token in moralis_data:
                balance = token.get('balance', '0')
                decimals = int(token.get('decimals', 18))
                symbol = token.get('symbol', 'UNKNOWN')
                
                # Convertir balance a decimal
                if balance and balance != '0':
                    balance_decimal = float(balance) / (10 ** decimals)
                    
                    if balance_decimal > 0.001:  # Solo balances significativos
                        positions.append({
                            'symbol': symbol,
                            'direction': 'LONG',
                            'entry_price': 0.0,
                            'size': balance_decimal,
                            'leverage': 1.0,
                            'platform': 'Moralis_API'
                        })
            
            return positions
            
        except Exception as e:
            logger.error(f"Error parseando datos Moralis: {e}")
            return []
    
    def load_advanced_wallet_config(self):
        """Carga las API keys desde advanced_wallet_config.json."""
        try:
            import os
            script_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
            config_path = os.path.join(script_dir, 'advanced_wallet_config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    api_keys = config.get('wallet_tracking', {}).get('api_keys', {})
                    logger.info(f"✅ Configuración avanzada cargada desde: {config_path}")
                    return api_keys
            else:
                logger.debug(f"Archivo de configuración avanzada no encontrado: {config_path}")
                return {}
                
        except Exception as e:
            logger.error(f"Error cargando configuración avanzada: {e}")
            return {}
    
    async def detect_positions_alternative_apis(self, wallet_address: str):
        """Detecta posiciones usando APIs alternativas más confiables."""
        try:
            wallet_positions = []
            
            # Método 1: Usar Alchemy API (más confiable que Etherscan)
            logger.info("🔍 Método 1: Consultando Alchemy API...")
            try:
                import aiohttp
                async with aiohttp.ClientSession() as session:
                    # Alchemy API para obtener balance de tokens
                    alchemy_url = f"https://eth-mainnet.g.alchemy.com/v2/demo/getTokenBalances"
                    params = {
                        'address': wallet_address,
                        'type': 'erc20'
                    }
                    
                    async with session.get(alchemy_url, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            positions = self.parse_alchemy_balances(data)
                            wallet_positions.extend(positions)
                            logger.info(f"✅ Alchemy API encontró {len(positions)} posiciones REALES")
                            for pos in positions:
                                logger.info(f"   📊 {pos['symbol']}: {pos['size']:.4f} ({pos['platform']})")
                        else:
                            logger.info(f"❌ Alchemy API error: {response.status}")
            except Exception as e:
                logger.info(f"❌ Error consultando Alchemy: {e}")
            
            # Método 2: Usar CoinGecko API para verificar posiciones
            if not wallet_positions:
                logger.info("🔍 Método 2: Consultando CoinGecko API...")
                try:
                    async with aiohttp.ClientSession() as session:
                        # API pública de CoinGecko (sin key requerida)
                        url = f"https://api.coingecko.com/api/v3/simple/price"
                        params = {
                            'ids': 'ethereum,bitcoin,tether,usd-coin',
                            'vs_currencies': 'usd'
                        }
                        
                        async with session.get(url, params=params) as response:
                            if response.status == 200:
                                # CoinGecko responde, pero no detecta posiciones específicas
                                logger.info("✅ CoinGecko API disponible, pero no detecta posiciones específicas de wallet")
                            else:
                                logger.info(f"❌ CoinGecko API error: {response.status}")
                except Exception as e:
                    logger.info(f"❌ Error consultando CoinGecko: {e}")
            
            # Método 3: Usar Zapper API (DeFi positions)
            if not wallet_positions:
                logger.info("🔍 Método 3: Consultando Zapper API...")
                try:
                    async with aiohttp.ClientSession() as session:
                        zapper_url = f"https://api.zapper.fi/v1/protocols/tokens/balances"
                        params = {
                            'addresses[]': wallet_address,
                            'network': 'ethereum'
                        }
                        
                        async with session.get(zapper_url, params=params) as response:
                            if response.status == 200:
                                data = await response.json()
                                positions = self.parse_zapper_balances(data)
                                wallet_positions.extend(positions)
                                logger.info(f"✅ Zapper API encontró {len(positions)} posiciones REALES")
                                for pos in positions:
                                    logger.info(f"   📊 {pos['symbol']}: {pos['size']:.4f} ({pos['platform']})")
                            else:
                                logger.info(f"❌ Zapper API error: {response.status}")
                except Exception as e:
                    logger.info(f"❌ Error consultando Zapper: {e}")
            
            # Resultado final
            if wallet_positions:
                logger.info(f"🎯 RESULTADO: {len(wallet_positions)} posiciones detectadas de APIs REALES")
                for i, pos in enumerate(wallet_positions, 1):
                    logger.info(f"   {i}. {pos['symbol']} {pos['direction']} - Tamaño: {pos['size']:.4f} - Fuente: {pos['platform']}")
            else:
                logger.info("🎯 RESULTADO: 0 posiciones detectadas - TODAS las APIs fallaron")
            
            return wallet_positions
            
        except Exception as e:
            logger.error(f"Error en APIs alternativas: {e}")
            return []
    
    def parse_alchemy_balances(self, alchemy_data):
        """Parsea datos de Alchemy para extraer posiciones."""
        try:
            positions = []
            
            if 'result' in alchemy_data and 'tokenBalances' in alchemy_data['result']:
                for token in alchemy_data['result']['tokenBalances']:
                    balance = token.get('tokenBalance', '0x0')
                    if balance and balance != '0x0':
                        # Convertir balance hexadecimal a decimal
                        balance_decimal = int(balance, 16) / 1e18
                        
                        if balance_decimal > 0.001:  # Solo balances significativos
                            # Mapear dirección de contrato a símbolo
                            contract_address = token.get('contractAddress', '').lower()
                            symbol = self.map_contract_to_symbol(contract_address)
                            
                            positions.append({
                                'symbol': symbol,
                                'direction': 'LONG',
                                'entry_price': 0.0,
                                'size': balance_decimal,
                                'leverage': 1.0,
                                'platform': 'Alchemy_Detected'
                            })
            
            return positions
            
        except Exception as e:
            logger.error(f"Error parseando datos Alchemy: {e}")
            return []
    
    def parse_zapper_balances(self, zapper_data):
        """Parsea datos de Zapper para extraer posiciones DeFi."""
        try:
            positions = []
            
            for address_data in zapper_data.values():
                for protocol_data in address_data.values():
                    for position in protocol_data.get('products', []):
                        for asset in position.get('assets', []):
                            balance = float(asset.get('balance', 0))
                            if balance > 0.001:
                                positions.append({
                                    'symbol': asset.get('symbol', 'UNKNOWN'),
                                    'direction': 'LONG',
                                    'entry_price': 0.0,
                                    'size': balance,
                                    'leverage': 1.0,
                                    'platform': 'Zapper_DeFi'
                                })
            
            return positions
            
        except Exception as e:
            logger.error(f"Error parseando datos Zapper: {e}")
            return []
    
    def map_contract_to_symbol(self, contract_address):
        """Mapea dirección de contrato a símbolo de token."""
        contract_mapping = {
            '0xdac17f958d2ee523a2206206994597c13d831ec7': 'USDT',
            '0xa0b86a33e6417c1c6c7c4b4c3d3c3c3c3c3c3c3c': 'USDC',
            '0x2260fac5e5542a773aa44fbcfedf7c193bc2c599': 'BTC',
            '0xc02aaa39b223fe8d0a0e5c4f27ead9083c756cc2': 'ETH',
            '0x0000000000000000000000000000000000000000': 'ETH',
        }
        
        return contract_mapping.get(contract_address, 'UNKNOWN')
    
    
    async def detect_positions_from_transactions(self, wallet_address: str):
        """Detecta posiciones analizando transacciones recientes de la wallet."""
        try:
            wallet_positions = []
            
            # Método 1: Usar API de Etherscan para transacciones recientes
            if hasattr(self, 'wallet_tracker') and self.wallet_tracker and self.wallet_tracker.etherscan_api_key:
                try:
                    import aiohttp
                    async with aiohttp.ClientSession() as session:
                        params = {
                            'module': 'account',
                            'action': 'txlist',
                            'address': wallet_address,
                            'startblock': 0,
                            'endblock': 99999999,
                            'page': 1,
                            'offset': 10,
                            'sort': 'desc',
                            'apikey': self.wallet_tracker.etherscan_api_key
                        }
                        
                        url = "https://api.etherscan.io/api"
                        async with session.get(url, params=params) as response:
                            if response.status == 200:
                                data = await response.json()
                                if data.get('status') == '1':
                                    # Analizar transacciones para detectar patrones de trading
                                    positions = self.analyze_trading_transactions(data.get('result', []))
                                    wallet_positions.extend(positions)
                except Exception as e:
                    logger.debug(f"Error analizando transacciones Etherscan: {e}")
            
            # Método 2: Usar API alternativas para detectar posiciones
            if not wallet_positions:
                try:
                    # Intentar con DeBank API si está disponible
                    if hasattr(self, 'wallet_tracker') and self.wallet_tracker and hasattr(self.wallet_tracker, 'debank_api_key'):
                        async with aiohttp.ClientSession() as session:
                            headers = {'AccessKey': self.wallet_tracker.debank_api_key} if self.wallet_tracker.debank_api_key else {}
                            url = f"https://pro-openapi.debank.com/v1/user/complex_protocol_list"
                            params = {'id': wallet_address}
                            
                            async with session.get(url, headers=headers, params=params) as response:
                                if response.status == 200:
                                    data = await response.json()
                                    positions = self.parse_debank_positions(data)
                                    wallet_positions.extend(positions)
                except Exception as e:
                    logger.debug(f"Error consultando DeBank para detección: {e}")
            
            logger.info(f"Posiciones detectadas mediante análisis: {len(wallet_positions)}")
            return wallet_positions
            
        except Exception as e:
            logger.error(f"Error en detección de posiciones por transacciones: {e}")
            return []
    
    def analyze_trading_transactions(self, transactions):
        """Analiza transacciones para identificar patrones de trading y posiciones activas."""
        try:
            positions = []
            
            # Contratos conocidos de plataformas de trading
            trading_contracts = {
                '0x65c7c7c4f3d6f5e4b4e4f4e4f4e4f4e4f4e4f4e4': 'dYdX',
                '0xe592427a0aece92de3edee1f18e0157c05861564': 'Uniswap_V3',
                '0x1111111254eeb25477b68fb85ed929f73a960582': '1inch',
                '0x3f5ce5fbfe3e9af3971dd833d26ba9b5c936f0be': 'Binance',
                # Agregar más contratos según sea necesario
            }
            
            # Analizar transacciones recientes (últimas 24 horas)
            from datetime import datetime, timedelta
            cutoff_time = datetime.now() - timedelta(hours=24)
            
            for tx in transactions[:20]:  # Analizar últimas 20 transacciones
                try:
                    tx_time = datetime.fromtimestamp(int(tx.get('timeStamp', 0)))
                    if tx_time < cutoff_time:
                        continue
                    
                    to_address = tx.get('to', '').lower()
                    value = float(tx.get('value', 0)) / 1e18
                    
                    # Verificar si es interacción con plataforma de trading
                    platform = trading_contracts.get(to_address)
                    if platform and value > 0.01:  # Transacciones significativas
                        # Inferir posición basada en la transacción
                        symbol = self.infer_symbol_from_transaction(tx)
                        if symbol:
                            positions.append({
                                'symbol': symbol,
                                'direction': 'LONG',  # Por defecto, se puede mejorar
                                'entry_price': 0.0,   # Se puede obtener del precio histórico
                                'size': value,
                                'leverage': 1.0,      # Se puede inferir mejor
                                'platform': platform
                            })
                            
                except Exception as e:
                    logger.debug(f"Error analizando transacción: {e}")
                    continue
            
            return positions
            
        except Exception as e:
            logger.error(f"Error analizando transacciones de trading: {e}")
            return []
    
    def infer_symbol_from_transaction(self, tx):
        """Infiere el símbolo de trading basado en los datos de la transacción."""
        try:
            # Mapeo básico de contratos a símbolos
            contract_to_symbol = {
                # ETH es el símbolo por defecto para transacciones ETH
                'eth': 'ETH',
                # Agregar más mapeos según sea necesario
            }
            
            # Si la transacción involucra ETH directamente
            value = float(tx.get('value', 0))
            if value > 0:
                return 'ETH'
            
            # Analizar input data para detectar otros tokens
            input_data = tx.get('input', '')
            if input_data and len(input_data) > 10:
                # Aquí se puede implementar lógica más sofisticada
                # para detectar otros símbolos basados en el input
                pass
            
            return 'ETH'  # Por defecto
            
        except Exception as e:
            logger.debug(f"Error infiriendo símbolo: {e}")
            return 'UNKNOWN'
    
    def parse_debank_positions(self, debank_data):
        """Parsea datos de DeBank para extraer posiciones activas."""
        try:
            positions = []
            
            for protocol in debank_data.get('data', []):
                protocol_name = protocol.get('name', 'Unknown')
                
                for portfolio in protocol.get('portfolio_item_list', []):
                    # Analizar posiciones de lending/borrowing
                    for asset in portfolio.get('detail', {}).get('supply_token_list', []):
                        symbol = asset.get('symbol', 'UNKNOWN')
                        amount = float(asset.get('amount', 0))
                        
                        if amount > 0.001:  # Solo posiciones significativas
                            positions.append({
                                'symbol': symbol,
                                'direction': 'LONG',
                                'entry_price': 0.0,
                                'size': amount,
                                'leverage': 1.0,
                                'platform': protocol_name
                            })
            
            return positions
            
        except Exception as e:
            logger.error(f"Error parseando posiciones DeBank: {e}")
            return []

# --- Función Main (sin cambios, usa cofigETHBTC.json) ---
async def main():
    script_dir = os.path.dirname(os.path.realpath(__file__))
    
    config_filename = 'cofigETHBTC.json' # <-- Usa tu config personalizado
    config_path = os.path.join(script_dir, config_filename)

    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
    except FileNotFoundError:
        logger.error(f"No se encontró '{config_filename}' en la ruta: {config_path}. Por favor, crea uno.")
        return

    bot = SmartMoneyLiveBot(
        api_key=config['api_key'],
        api_secret=config['api_secret'],
        telegram_token=config['telegram_token'],
        telegram_chat_id=config['telegram_chat_id'],
        symbols=config.get('symbols', ['BTCUSDT']),
        initial_balance=float(config.get('initial_balance', 1000.0))
    )
    # Aplicar parámetros opcionales desde JSON
    try:
        if 'enable_macd_filter' in config:
            bot.enable_macd_filter = bool(config['enable_macd_filter'])
        if 'macd_fast' in config:
            bot.macd_fast = int(config['macd_fast'])
        if 'macd_slow' in config:
            bot.macd_slow = int(config['macd_slow'])
        if 'macd_signal' in config:
            bot.macd_signal = int(config['macd_signal'])
        if 'max_concurrent_open' in config:
            bot.max_concurrent_open = int(config['max_concurrent_open'])
        if 'leverage_per_symbol' in config and isinstance(config['leverage_per_symbol'], dict):
            # Normalizar claves a str sin barras
            bot.leverage_per_symbol = {str(k): int(v) for k, v in config['leverage_per_symbol'].items()}
        
        # Cargar configuraciones específicas por símbolo
        if 'symbol_configs' in config and isinstance(config['symbol_configs'], dict):
            bot.symbol_configs = config['symbol_configs']
            logger.info(f"✅ Configuraciones por símbolo cargadas para: {', '.join(bot.symbol_configs.keys())}")
            for symbol, cfg in bot.symbol_configs.items():
                logger.info(f"   {symbol}: TP={cfg.get('tp_percentage')}%, SL={cfg.get('sl_percentage')}%, "
                          f"Structural Stop={'✅' if cfg.get('enable_structural_stop') else '❌'}, "
                          f"R:R={cfg.get('risk_reward_ratio')}")
    except Exception as e:
        logger.error(f"Error aplicando configuración opcional desde JSON: {e}")
    
    # Inicializar wallet tracking si está habilitado
    wallet_config = config.get('wallet_tracking', {})
    if wallet_config.get('enabled', False):
        wallet_address = wallet_config.get('wallet_address')
        if wallet_address:
            etherscan_key = wallet_config.get('etherscan_api_key') or None
            moralis_key = wallet_config.get('moralis_api_key') or None
            debank_key = wallet_config.get('debank_api_key') or None
            copy_trading = wallet_config.get('copy_trading', True)
            
            logger.info(f"🔍 Configurando rastreo de wallet: {wallet_address}")
            logger.info(f"🔄 Copy trading: {'✅ HABILITADO' if copy_trading else '❌ DESHABILITADO'}")
            
            # Log de inicio del rastreo
            logger.info("=" * 60)
            logger.info("🚀 INICIANDO WALLET TRACKER CON COPY TRADING")
            logger.info("=" * 60)
            
            # Iniciar wallet tracking después de un pequeño delay
            asyncio.create_task(bot.start_wallet_tracking(wallet_address, etherscan_key, moralis_key, debank_key, copy_trading))
        else:
            logger.warning("Wallet tracking habilitado pero no se especificó dirección")
    
    try:
        await bot.run()
    except KeyboardInterrupt:
        logger.info("Deteniendo bot...")
        await bot.send_telegram_message("🛑 <b>Bot Detenido Manualmente</b>")
        bot.stop()

if __name__ == "__main__":
    asyncio.run(main())